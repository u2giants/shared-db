#!/usr/bin/env python3
"""Idempotent importer for the legacy Google OrderList `Order` tab.

Issue:   #727 (PopDAM OrderList Step 3)
Plan:    plan_popdam_order_list.md, Phase 2 Step 2.2
Profile: docs/app-migration-notes/popdam-order-list.md (the 48-column contract)
Schema:  supabase/migrations/20260810010000_popdam_order_list_contract.sql

=====================================================================================
READ THIS FIRST — WHERE THIS IMPORT HAS AND HAS NOT RUN
=====================================================================================
PREVIEW (``rjyboqwcdzcocqgmsyel``): DONE, 2026-08-12. 3,212 orders and 24,010 lines
inserted; a second identical run changed 0 business rows; all 10 balance checks PASS.
Evidence: ``docs/verification/popdam-order-list-preview-2026-08-12/README.md``.

PRODUCTION (``qsllyeztdwjgirsysgai``): NOT RUN. NOTHING has been written. The
``--production`` path added for issue #852 exists so an EXACT, commit-bound approval
package can be written; it has never been executed, and it must not be executed until
the owner approves the exact command and confirmation. A green CI run on this file
proves the LOGIC and the REFUSALS, not an import. Do not read one as the other.

SOURCE ROW COUNT — SETTLED BY OWNER RULING. The approved populated-row count is
**12,354**, for the approved ``68c9b03a…`` workbook under this file's own
``is_populated`` definition. Two earlier figures appear in the history (12,328 from the
superseded 2026-08-09 export, and 12,323 from 2026-08-11 live-sheet exports that were
never retained); the owner closed that question by decision and no itemisation of the
five-row difference is required. The history is kept in the reconciliation report as
provenance, not as an open gate.

``--expected-populated-rows`` remains required and un-defaulted anyway. That is not
because the number is in doubt — it is so the approved figure is stated explicitly in
the command being approved, and asserted against the workbook before any write, rather
than living only as a constant in this file.

=====================================================================================
SAFETY RULES THIS FILE ENFORCES
=====================================================================================
* The workbook SHA-256 must equal ``--expected-sha256`` (defaulting to the approved
  constant below). A different workbook is refused, not imported "close enough".
* Writes require an EXPLICIT mode flag AND ``supabase/.temp/project-ref`` reading the
  matching ref: ``--preview`` requires ``rjyboqwcdzcocqgmsyel``; ``--production``
  requires ``qsllyeztdwjgirsysgai``. The two flags are mutually exclusive. Neither
  flag means no write at all.
* ``--production`` additionally requires (issue #852, and the approval package in
  ``docs/verification/popdam-order-list-production-approval-2026-08-12/README.md``):
  the exact approved workbook SHA-256 (``--expected-sha256`` cannot relax it);
  ``--reviewed-commit`` naming the merged, reviewed shared-db commit; ``--confirm``
  carrying the exact confirmation sentence bound to that commit, that hash and that
  project ref; and ``--expected-populated-rows`` stating the count the operator
  expects. Every one of those is checked BEFORE the connection string is read, so a
  wrong invocation never touches a credential.
* The write target is re-proved from ``supabase/.temp/project-ref`` immediately before
  the connection is opened AND again before every batch, so a relink mid-run aborts
  instead of writing the remaining batches into the wrong database.
* ``--replace-source`` is impossible in production. It is blocked by the CLI, and
  ``apply_plan`` is additionally called with ``allow_replace=False``, which raises
  before any write even if the CLI check were ever removed.
* Every canonical row is addressed by a deterministic Google source ref. A second
  identical run therefore inserts nothing and updates nothing.
* An existing row whose desired payload has DRIFTED is never silently rewritten. It is
  counted and reported, and only ``--replace-source`` (preview-only) updates it.
* No raw customer or order text is written to the report or to stdout. Only counts and
  deterministic source refs, which contain a sheet row number and a PO number that the
  profile document already publishes in aggregate.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

# =====================================================================================
# Source identity. Deliberately module-level and readable, not buried in argparse
# defaults, so a reviewer can diff the approved checksum in one glance.
# =====================================================================================
SOURCE_SPREADSHEET_ID = "1i1da5J0qy5a0EvsO1CvfyQ6Xijn4678LG7TFqbwxwUk"
SOURCE_TAB_NAME = "Order"
SOURCE_TAB_GID = 0
SOURCE_SYSTEM = "google_order_list"

#: SHA-256 of the owner-accepted 2026-08-11 read-only export
#: (``C:\\Users\\ahazan2\\Downloads\\OrderList.xlsx``), ruled authoritative by Albert on
#: 2026-08-11 (issue #727, "accept today's sheet"), superseding the 2026-08-09 export
#: (``4958b4b7...``). Re-profiled honestly on 2026-08-12; see
#: docs/app-migration-notes/popdam-order-list.md "Re-profile 2026-08-12".
#:
#: Column-AR note (root-caused, fixed): this workbook fills column AR ("Sub SKU") with a
#: copy of column P ("Style#") on every direct-order row. That is a spreadsheet artifact,
#: not an assortment component list, so sub_sku_mirrors_style() excludes an AR that
#: exactly equals the Style# and carries no newline from the assortment signal. Genuine
#: multiline assortments are unaffected. The baseline below is derived from that corrected
#: logic; the eight conditional assertions stay armed against the new SHA.
APPROVED_SOURCE_SHA256 = (
    "68c9b03a0ec183e08b3a8f2344397e1bc4f61e73457849e7bf8c0cf7fb2409fe"
)

PREVIEW_PROJECT_REF = "rjyboqwcdzcocqgmsyel"
PRODUCTION_PROJECT_REF = "qsllyeztdwjgirsysgai"

DEFAULT_PROJECT_REF_FILE = "supabase/.temp/project-ref"
DEFAULT_BATCH_SIZE = 500

#: Run modes. ``dry_run`` writes nothing at all; the other two write.
MODE_DRY_RUN = "dry-run"
MODE_PREVIEW = "preview"
MODE_PRODUCTION = "production"

#: The project ref each write mode is allowed to touch. Any other value refuses.
MODE_REQUIRED_PROJECT_REF = {
    MODE_PREVIEW: PREVIEW_PROJECT_REF,
    MODE_PRODUCTION: PRODUCTION_PROJECT_REF,
}

#: The exact sentence ``--confirm`` must carry for a production run. It is bound to the
#: reviewed commit, the workbook hash and the destination project ref, so a confirmation
#: copied from a different approval package, a different workbook, or a different target
#: cannot authorize this one. The approval package quotes this same wording.
PRODUCTION_CONFIRMATION_TEMPLATE = (
    "I approve one production import of OrderList.xlsx "
    "SHA-256 {source_sha256} "
    "into Supabase project {project_ref} "
    "using reviewed shared-db commit {commit_sha}"
)

#: Length of a full git object name. Production refuses an abbreviated SHA: an approval
#: package that names seven characters is not an exact package.
FULL_GIT_SHA_LENGTH = 40

# =====================================================================================
# Server identity. Reading supabase/.temp/project-ref proves the Supabase CLI's LINK
# state; it says nothing about the database the connection string actually opens. The
# two are only connected if we ask the SERVER. Supabase does not expose the project ref
# through SQL, so the ref is recovered from the connection parameters libpq ACTUALLY
# used (connection.info), which is the socket that carries the writes — not from the
# string we parsed. That is combined with a live content precondition that a
# preview-pointing URL cannot satisfy.
# =====================================================================================
#: ``db.<ref>.supabase.co`` (direct) and ``<ref>.supabase.co``.
HOST_PROJECT_REF_PATTERNS = (
    re.compile(r"^db\.([a-z0-9]{20})\.supabase\.(?:co|com|net|in)$", re.IGNORECASE),
    re.compile(r"^([a-z0-9]{20})\.supabase\.(?:co|com|net|in)$", re.IGNORECASE),
)
#: Pooler connections carry the ref in the user name: ``postgres.<ref>``.
POOLER_USER_REF_PATTERN = re.compile(r"^postgres\.([a-z0-9]{20})$", re.IGNORECASE)

#: Supabase's TRANSACTION-mode pooler. It cannot hold a multi-statement transaction, so
#: the single-transaction production import would be silently split. Refused.
TRANSACTION_POOLER_PORT = 6543

#: Advisory-lock key for the whole OrderList import. Two overlapping runs would both
#: pass the read-then-insert check in ``order_id_for_source`` and both insert.
IMPORT_ADVISORY_LOCK_KEY = 852_000_727


# =====================================================================================
# Phase 0 baseline. The reconciliation report asserts against these numbers; it does
# not merely print whatever the run happened to produce.
# =====================================================================================
@dataclass(frozen=True)
class ReconciliationBaseline:
    """Counts proven by the Phase 0 profile, for the approved workbook only.

    Re-derived honestly on 2026-08-12 from the owner-accepted 2026-08-11 export by
    running build_plan() over the real ``Order`` tab with an empty catalog through the
    corrected classify_row_shape()/sub_sku_mirrors_style() logic (see the private profiler
    artifact under the run's temp dir). These are the TRUE counts of that workbook.

    The workbook mirrors the Style# into column AR on direct rows; sub_sku_mirrors_style()
    excludes that artifact from the assortment signal, so the 8,438 direct rows classify as
    ROW_SHAPE_DIRECT (only the 3 rows that ALSO carry an Assortment ID in column O remain
    both-shape). Under the pre-fix logic these read as direct_only=0 / both_shape=8,441.
    """

    populated_rows: int = 12_354
    direct_only_rows: int = 8_438
    assortment_only_rows: int = 3_899
    both_shape_rows: int = 3
    neither_shape_rows: int = 14
    assortment_components: int = 15_713
    structurally_invalid_assortment_rows: int = 28
    #: Catalog-dependent, NOT derivable from the workbook alone (needs live Master Data
    #: duplicate SKUs). Re-derived at import time against the real catalog; the 2026-08-09
    #: profile counted 449. Not one of the eight conditional assertions and not used in
    #: row_shapes_balance(); left as a documented prior figure, re-confirm on the real run.
    ambiguous_matches: int = 449
    normalized_po_numbers: int = 3_087
    blank_po_rows: int = 138

    def row_shapes_balance(self) -> bool:
        return self.populated_rows == (
            self.direct_only_rows
            + self.assortment_only_rows
            + self.both_shape_rows
            + self.neither_shape_rows
        )


BASELINE = ReconciliationBaseline()


# =====================================================================================
# The 48-column map. Letters, not indexes, because the profile document and the sheet
# both speak in letters and an off-by-one here silently imports the wrong field.
# =====================================================================================
COL_PO_STATUS = "A"
COL_IMPORT_PO = "B"
COL_ORDER_VENDOR = "C"
COL_SEAL_CONTAINER_DAY = "D"
COL_SENT_PO_DATE = "E"
COL_DEFAULT_VENDOR = "F"
COL_SAMPLED_VENDOR = "G"
COL_VENDOR_ID = "H"
COL_COMPANY = "I"
COL_ORDER_PERSON = "J"
COL_ORDER_TYPE = "K"
COL_CUSTOMER = "L"
COL_SUFFIX = "M"
COL_CUSTOMER_PO = "N"
COL_ASSORTMENT_ID = "O"
COL_STYLE = "P"
COL_SAMPLE_DEPTH = "Q"
COL_ORDER_DEPTH = "R"
COL_DESCRIPTION = "S"
COL_LICENSE_STATUS = "T"
COL_QUANTITY = "U"
COL_CASE_PACK = "V"
COL_CASES = "W"
COL_SHIP_TO = "X"
COL_START_SHIP = "Y"
COL_CANCEL_DATE = "Z"
COL_VENDOR_DELIVERY = "AA"
COL_CARGO_FORECAST = "AB"
COL_BOOK_CONTAINER = "AC"
COL_ETD = "AD"
COL_ETA = "AE"
COL_TEST_REPORT = "AF"
COL_PROFESSIONAL_PHOTOS = "AG"
COL_WAREHOUSE_DAY = "AH"
COL_CONTAINER_BOOKING = "AI"
COL_MBL = "AJ"
COL_CLOSE_TRACKING = "AK"
COL_NEW_REPEAT = "AL"
COL_LICENSOR_OR_GENERIC = "AM"
COL_CONTRACTUAL_SAMPLE = "AN"
COL_BLANK_AO = "AO"
COL_FOR_WRITING_PO = "AP"
COL_ORDER_TAB_LINE = "AQ"
COL_SUB_SKU = "AR"
COL_SUB_QTY = "AS"
COL_SUB_DESCRIPTION = "AT"
COL_BLANK_AU = "AU"
COL_BLANK_AV = "AV"

LAST_COLUMN = COL_BLANK_AV
EXPECTED_COLUMN_COUNT = 48

#: Columns whose value must be IDENTICAL across every row of one normalized Import PO.
#: A conflict quarantines the whole PO group. The profile counted five vendor, five
#: customer, one status and one company conflict; choosing the first or most frequent
#: value is explicitly forbidden (plan section 8, open judgment 2).
HEADER_IDENTITY_COLUMNS = (
    COL_PO_STATUS,
    COL_ORDER_VENDOR,
    COL_VENDOR_ID,
    COL_COMPANY,
    COL_CUSTOMER,
)

#: Header-scoped payload columns. A disagreement here does NOT quarantine the order —
#: the profile proved these are stable "in nearly every" group — but the typed value is
#: left NULL and the disagreement is counted, never averaged or first-wins.
HEADER_PAYLOAD_COLUMNS = (
    COL_SEAL_CONTAINER_DAY,
    COL_SENT_PO_DATE,
    COL_VENDOR_DELIVERY,
    COL_BOOK_CONTAINER,
    COL_ETD,
    COL_ETA,
    COL_WAREHOUSE_DAY,
    COL_CONTAINER_BOOKING,
    COL_MBL,
    COL_CLOSE_TRACKING,
)

#: Newline-aligned assortment component lists. AR carries the component SKUs and sets
#: the expected count; every other list must match it exactly or the row is rejected.
ASSORTMENT_ALIGNED_COLUMNS = (
    COL_SUB_SKU,
    COL_LICENSOR_OR_GENERIC,
    COL_SUB_DESCRIPTION,
    COL_LICENSE_STATUS,
    COL_TEST_REPORT,
    COL_PROFESSIONAL_PHOTOS,
)

#: Values the sheet uses for "no value". They are evidence, never data.
SPREADSHEET_ERROR_VALUES = frozenset(
    {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}
)

TRUE_MARKERS = frozenset({"yes", "y", "true", "1", "x", "done", "ok", "booked"})
FALSE_MARKERS = frozenset({"no", "n", "false", "0"})

ROW_SHAPE_DIRECT = "direct"
ROW_SHAPE_ASSORTMENT = "assortment"
ROW_SHAPE_BOTH = "both"
ROW_SHAPE_NEITHER = "neither"

MATCH_MATCHED = "matched"
MATCH_UNMATCHED = "unmatched"
MATCH_AMBIGUOUS = "ambiguous"
MATCH_MANUAL = "manual"
MATCH_NOT_APPLICABLE = "not_applicable"

RESOLUTION_UNIQUE = "unique"
RESOLUTION_AMBIGUOUS = "ambiguous"
RESOLUTION_UNMATCHED = "unmatched"
RESOLUTION_NO_TYPE = "no_type"


class ImporterError(RuntimeError):
    """Every refusal in this file. Loud, never swallowed."""


class CommitOutcomeUnknown(ImporterError):
    """The single COMMIT was issued but its outcome could not be established.

    This is NOT the same as a rollback and must never be reported as one. If the
    connection drops between the server committing and the acknowledgement arriving,
    the rows ARE durable and the client cannot tell. Data integrity still holds — a
    rerun meets :func:`assert_no_existing_source_refs`, finds the rows and refuses —
    but the durable report must say "unknown, go and look", not "nothing was written".
    Claiming a fact the code cannot know is the same defect class as a false record.
    """


# =====================================================================================
# Column-letter arithmetic
# =====================================================================================
def column_index(letter: str) -> int:
    """Zero-based index for a spreadsheet column letter. ``A`` -> 0, ``AV`` -> 47."""
    letter = letter.strip().upper()
    if not letter or not letter.isalpha():
        raise ImporterError(f"Not a column letter: {letter!r}")
    index = 0
    for char in letter:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def column_letter(index: int) -> str:
    """Inverse of :func:`column_index`."""
    if index < 0:
        raise ImporterError(f"Negative column index: {index}")
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


# =====================================================================================
# Normalization. Trim + case-fold ONLY, matching
# plm.production_order_line.sku_normalized. Punctuation stripping is forbidden: it is
# how two different products silently become one.
# =====================================================================================
def normalize_text(value: Any) -> str | None:
    """Trim to a string, collapse blanks and spreadsheet errors to ``None``."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, (datetime, date)):
        text = value.isoformat()
    else:
        text = str(value).strip()
    if not text:
        return None
    if text.upper() in SPREADSHEET_ERROR_VALUES:
        return None
    return text


def raw_text(value: Any) -> str | None:
    """Trim only. Preserves ``#N/A`` and friends, because they are source evidence."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, (datetime, date)):
        text = value.isoformat()
    else:
        text = str(value).strip()
    return text or None


def normalize_sku(value: Any) -> str | None:
    """``nullif(btrim(lower(sku)), '')`` — byte-identical to the generated column."""
    text = normalize_text(value)
    return text.lower() if text else None


def normalize_po_number(value: Any) -> str | None:
    """Trim + case-fold. 3,088 raw PO values collapse to 3,083 normalized ones."""
    text = normalize_text(value)
    return text.lower() if text else None


def normalize_style_type(value: Any) -> str | None:
    """Column AM to ``licensed`` / ``generic``. Anything else is ``None``, not a guess.

    Five cells hold ``#N/A`` and 24 are blank; both become ``None`` and the resulting
    line carries ``master_data_match_status = 'not_applicable'`` rather than being
    matched against an arbitrarily chosen catalog.
    """
    text = normalize_text(value)
    if text is None:
        return None
    lowered = text.lower()
    if lowered == "generic":
        return "generic"
    # The sheet writes the licensor's NAME here for licensed styles ("Disney",
    # "Warner"), not the word "licensed". Anything that is not exactly "generic" and is
    # not blank/error is therefore a licensed style.
    return "licensed"


_ISO_DATE = re.compile(r"^(\d{4})-(\d{1,2})-(\d{1,2})$")
_US_DATE = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})$")


def parse_date(value: Any) -> tuple[date | None, str | None]:
    """Return ``(typed_date, raw_text)``.

    Invalid values (``ASAP``, ``1//12/2022``, ``CANCELLED``, ``NODATE``, the two
    impossible ETD serials) yield ``(None, raw)``. The raw string is kept as evidence
    and the typed column stays NULL. Nothing is ever coerced into a plausible date.
    """
    raw = raw_text(value)
    if raw is None:
        return None, None
    if isinstance(value, datetime):
        return value.date(), raw
    if isinstance(value, date):
        return value, raw
    if raw.upper() in SPREADSHEET_ERROR_VALUES:
        return None, raw

    match = _ISO_DATE.match(raw)
    if match:
        year, month, day = (int(part) for part in match.groups())
        return _safe_date(year, month, day), raw

    match = _US_DATE.match(raw)
    if match:
        month, day, year_text = match.group(1), match.group(2), match.group(3)
        year = int(year_text)
        if len(year_text) == 2:
            year += 2000 if year < 70 else 1900
        return _safe_date(year, int(month), int(day)), raw

    # Bare integers are Excel serial dates. The profile found two impossible ones
    # (6693547); a plausibility window rejects them instead of storing year 20,000.
    if re.fullmatch(r"\d+(\.0+)?", raw):
        serial = int(float(raw))
        if 1 <= serial <= 80_000:  # 80000 ~= year 2119. Anything beyond is corrupt.
            return date.fromordinal(date(1899, 12, 30).toordinal() + serial), raw
        return None, raw

    return None, raw


def _safe_date(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_number(value: Any) -> tuple[Decimal | None, str | None]:
    """Return ``(typed_number, raw_text)``.

    Column W holds 531 ``Wrong QTY`` strings; column V holds 380 text/multiline values.
    Those stay raw and the typed column stays NULL. Never compute over a coerced value.
    """
    raw = raw_text(value)
    if raw is None:
        return None, None
    if isinstance(value, bool):
        return None, raw
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)), raw
    if raw.upper() in SPREADSHEET_ERROR_VALUES:
        return None, raw
    cleaned = raw.replace(",", "").replace("$", "").strip()
    try:
        return Decimal(cleaned), raw
    except (InvalidOperation, ValueError):
        return None, raw


def parse_marker(value: Any) -> tuple[bool | None, str | None]:
    """Return ``(typed_boolean, raw_text)`` from an exact populated marker only."""
    raw = raw_text(value)
    if raw is None:
        return None, None
    lowered = raw.strip().lower()
    if lowered in TRUE_MARKERS:
        return True, raw
    if lowered in FALSE_MARKERS:
        return False, raw
    return None, raw


def split_aligned(value: Any) -> list[str | None]:
    """Split a newline-aligned assortment helper cell, preserving blank positions."""
    raw = raw_text(value)
    if raw is None:
        return []
    parts = re.split(r"\r\n|\r|\n", raw)
    return [normalize_text(part) for part in parts]


# =====================================================================================
# Deterministic Google source references
#
# Column AQ ("Order Tab Line#") is unusable — one #REF! and 12,327 blanks — so every id
# is GENERATED from the physical sheet row, which is what makes a repeat run a no-op.
# =====================================================================================
def header_source_id(po_normalized: str | None, sheet_row: int) -> str:
    """``order:po:<po>`` for a real PO; ``order:row:<row>`` for the 130 blank-PO rows.

    Blank-PO rows must NEVER collapse into one shared header, so each gets its own.
    """
    if po_normalized:
        return f"order:po:{po_normalized}"
    return f"order:row:{sheet_row}"


def line_source_id(sheet_row: int) -> str:
    return f"order:row:{sheet_row}"


def component_source_id(sheet_row: int, ordinal: int) -> str:
    if ordinal < 1:
        raise ImporterError(f"Component ordinal is 1-based, got {ordinal}")
    return f"order:row:{sheet_row}:component:{ordinal}"


# =====================================================================================
# Source rows
# =====================================================================================
@dataclass(frozen=True)
class SourceRow:
    """One physical row of the Google ``Order`` tab, untouched."""

    sheet_row: int
    values: tuple[Any, ...]

    def cell(self, letter: str) -> Any:
        index = column_index(letter)
        if index >= len(self.values):
            return None
        return self.values[index]

    def text(self, letter: str) -> str | None:
        return normalize_text(self.cell(letter))

    def raw(self, letter: str) -> str | None:
        return raw_text(self.cell(letter))

    def is_populated(self) -> bool:
        """A row counts as populated when any mapped cell holds anything at all."""
        return any(raw_text(value) is not None for value in self.values)


_COMPONENT_SEPARATOR = re.compile(r"[\r\n]")


def sub_sku_mirrors_style(row: SourceRow) -> bool:
    """True when column AR is just a copy of the direct Style# (column P), not a list.

    The owner-accepted 2026-08-11 workbook (issue #727) fills column AR ("Sub SKU") with
    the row's own Style# on every direct-order row. That is a spreadsheet artifact, not an
    assortment component list, so it must NOT be read as an assortment signal — otherwise
    every direct row would classify as ROW_SHAPE_BOTH and be rejected.

    The mirror is recognised narrowly, so a genuine assortment is never mistaken for one:

    * AR must be present and hold NO component separator (a real assortment list is
      newline-separated; any ``\\r``/``\\n`` disqualifies the shortcut), and
    * AR must equal the direct Style# under the importer's own SKU normalization
      (``normalize_sku`` = trim + case-fold, byte-identical to ``sku_normalized``).

    A multiline AR (even one whose first line equals the Style#), an AR that differs from
    the Style#, or an AR on a row with no direct Style# is left as a true assortment
    signal. The raw AR text is preserved as evidence by the callers; nothing is discarded.
    """
    ar_raw = row.raw(COL_SUB_SKU)
    if ar_raw is None or _COMPONENT_SEPARATOR.search(ar_raw):
        return False
    style_normalized = normalize_sku(row.cell(COL_STYLE))
    sub_sku_normalized = normalize_sku(row.cell(COL_SUB_SKU))
    return (
        style_normalized is not None
        and sub_sku_normalized is not None
        and style_normalized == sub_sku_normalized
    )


def classify_row_shape(row: SourceRow) -> str:
    """Direct SKU row, assortment row, both (quarantine) or neither (reject)."""
    has_direct = row.text(COL_STYLE) is not None
    # Column AR only signals an assortment when it is a genuine component list, not a
    # Style# mirror (see sub_sku_mirrors_style). Column O (Assortment ID) still counts.
    has_component_list = (
        row.text(COL_SUB_SKU) is not None and not sub_sku_mirrors_style(row)
    )
    has_assortment = has_component_list or row.text(COL_ASSORTMENT_ID) is not None
    if has_direct and has_assortment:
        return ROW_SHAPE_BOTH
    if has_direct:
        return ROW_SHAPE_DIRECT
    if has_assortment:
        return ROW_SHAPE_ASSORTMENT
    return ROW_SHAPE_NEITHER


@dataclass(frozen=True)
class AssortmentComponent:
    ordinal: int
    sku: str | None
    style_type: str | None
    description: str | None
    license_status: str | None
    test_report: str | None
    professional_photos: str | None


@dataclass(frozen=True)
class AssortmentSplit:
    components: tuple[AssortmentComponent, ...]
    rejected: bool
    reason: str | None
    #: ``{column_letter: observed_count}`` for every misaligned list. Report evidence.
    counts: Mapping[str, int] = field(default_factory=dict)


def split_assortment(row: SourceRow) -> AssortmentSplit:
    """Expand a Google assortment row into aligned component candidates.

    Alignment is judged against column AR (component SKUs). Any other populated helper
    list with a different length rejects the whole row: the profile found 10 such rows
    and mis-pairing a SKU with the wrong licensor is exactly the silent corruption the
    plan forbids. Component QUANTITY is never derived — column AS is entirely blank and
    guessing it is prohibited.
    """
    skus = split_aligned(row.cell(COL_SUB_SKU))
    if not skus:
        return AssortmentSplit((), True, "no_component_skus", {})

    expected = len(skus)
    counts: dict[str, int] = {COL_SUB_SKU: expected}
    aligned: dict[str, list[str | None]] = {COL_SUB_SKU: skus}

    for letter in ASSORTMENT_ALIGNED_COLUMNS:
        if letter == COL_SUB_SKU:
            continue
        parts = split_aligned(row.cell(letter))
        if not parts:
            aligned[letter] = [None] * expected
            continue
        counts[letter] = len(parts)
        if len(parts) != expected:
            return AssortmentSplit((), True, f"component_count_mismatch:{letter}", counts)
        aligned[letter] = parts

    # The discriminator list is what selects the eligible Master Data catalog. A row
    # whose AR list is aligned but whose AM cell was a single #N/A has no discriminator
    # at all; the components are still created, but as not_applicable evidence.
    components = tuple(
        AssortmentComponent(
            ordinal=ordinal,
            sku=aligned[COL_SUB_SKU][ordinal - 1],
            style_type=normalize_style_type(aligned[COL_LICENSOR_OR_GENERIC][ordinal - 1]),
            description=aligned[COL_SUB_DESCRIPTION][ordinal - 1],
            license_status=aligned[COL_LICENSE_STATUS][ordinal - 1],
            test_report=aligned[COL_TEST_REPORT][ordinal - 1],
            professional_photos=aligned[COL_PROFESSIONAL_PHOTOS][ordinal - 1],
        )
        for ordinal in range(1, expected + 1)
    )

    if all(component.sku is None for component in components):
        return AssortmentSplit((), True, "all_component_skus_blank", counts)

    return AssortmentSplit(components, False, None, counts)


# =====================================================================================
# Master Data resolution
# =====================================================================================
@dataclass(frozen=True)
class MasterDataCatalog:
    """``(normalized_sku, style_type)`` to the Master Data rows that claim it.

    ``item_ids`` maps a Master Data row id to its canonical ``plm.item`` id through
    ``plm.style_tracker_item_bridge``. TODAY THAT MAP IS EMPTY, because ``plm.item`` is
    unpopulated pending fix_schema_for_api.md Phase 4. That is not an error: the
    importer records ``master_data_resolution = 'unique'`` as evidence and leaves both
    ``item_id`` and ``master_data_match_status = 'matched'`` alone, exactly as issue
    #727's coordination constraint requires.
    """

    rows: Mapping[tuple[str, str], Sequence[str]] = field(default_factory=dict)
    item_ids: Mapping[str, str | None] = field(default_factory=dict)

    def candidates(self, sku: str | None, style_type: str | None) -> Sequence[str]:
        if not sku or not style_type:
            return ()
        return tuple(self.rows.get((sku, style_type), ()))


@dataclass(frozen=True)
class MatchResult:
    #: Value for plm.production_order_line.master_data_match_status.
    status: str
    #: plm.item.id, or None when the canonical link cannot yet be established.
    item_id: str | None
    #: Master Data evidence, independent of whether plm.item is populated yet.
    resolution: str
    #: Why item_id is null despite a unique Master Data row. Report + metadata only.
    pending_reason: str | None = None
    candidate_count: int = 0


def resolve_match(
    sku: str | None, style_type: str | None, catalog: MasterDataCatalog
) -> MatchResult:
    """Exact unique normalized SKU + type only. Never fuzzy, never first-of-duplicates."""
    if not sku:
        return MatchResult(MATCH_NOT_APPLICABLE, None, RESOLUTION_NO_TYPE, "no_sku", 0)
    if not style_type:
        return MatchResult(
            MATCH_NOT_APPLICABLE, None, RESOLUTION_NO_TYPE, "no_style_type", 0
        )

    candidates = catalog.candidates(sku, style_type)
    if not candidates:
        return MatchResult(MATCH_UNMATCHED, None, RESOLUTION_UNMATCHED, None, 0)
    if len(candidates) > 1:
        # 449 of these exist because Master Data itself holds duplicate normalized
        # SKUs. Visible data quality, NOT permission to pick the first row.
        return MatchResult(
            MATCH_AMBIGUOUS, None, RESOLUTION_AMBIGUOUS, None, len(candidates)
        )

    item_id = catalog.item_ids.get(candidates[0])
    if item_id is None:
        return MatchResult(
            MATCH_UNMATCHED, None, RESOLUTION_UNIQUE, "plm_item_unpopulated", 1
        )
    return MatchResult(MATCH_MATCHED, item_id, RESOLUTION_UNIQUE, None, 1)


# =====================================================================================
# The import plan: pure, deterministic, database-free
# =====================================================================================
@dataclass
class PlannedOrder:
    source_id: str
    po_normalized: str | None
    payload: dict[str, Any]
    metadata: dict[str, Any]
    sheet_rows: list[int]
    quarantined: bool = False
    quarantine_reasons: list[str] = field(default_factory=list)


@dataclass
class PlannedLine:
    source_id: str
    order_source_id: str
    sheet_row: int
    ordinal: int | None
    payload: dict[str, Any]
    metadata: dict[str, Any]


@dataclass
class RejectedRow:
    sheet_row: int
    shape: str
    reason: str
    evidence: dict[str, Any]


@dataclass
class ImportPlan:
    orders: list[PlannedOrder] = field(default_factory=list)
    lines: list[PlannedLine] = field(default_factory=list)
    rejected_rows: list[RejectedRow] = field(default_factory=list)
    counters: dict[str, int] = field(default_factory=dict)

    def bump(self, key: str, amount: int = 1) -> None:
        self.counters[key] = self.counters.get(key, 0) + amount

    def writable_orders(self) -> list[PlannedOrder]:
        return [order for order in self.orders if not order.quarantined]

    def writable_lines(self) -> list[PlannedLine]:
        blocked = {o.source_id for o in self.orders if o.quarantined}
        return [line for line in self.lines if line.order_source_id not in blocked]


def _single_value(values: Iterable[str | None]) -> tuple[str | None, bool]:
    """Return ``(value, conflicted)`` for a header field across one PO group."""
    distinct = {value for value in values if value is not None}
    if not distinct:
        return None, False
    if len(distinct) > 1:
        return None, True
    return distinct.pop(), False


def build_plan(rows: Sequence[SourceRow], catalog: MasterDataCatalog) -> ImportPlan:
    """Turn staged source rows into canonical order/line candidates.

    This function touches no database and performs no I/O. Everything the reconciliation
    report claims is derived here, which is why the whole contract is unit-testable.
    """
    plan = ImportPlan()
    populated = [row for row in rows if row.is_populated()]
    plan.bump("staged_rows", len(populated))

    # ---- 1. Group rows into headers -------------------------------------------------
    groups: dict[str, list[SourceRow]] = {}
    group_po: dict[str, str | None] = {}
    for row in populated:
        po_normalized = normalize_po_number(row.cell(COL_IMPORT_PO))
        source_id = header_source_id(po_normalized, row.sheet_row)
        groups.setdefault(source_id, []).append(row)
        group_po[source_id] = po_normalized
        if po_normalized is None:
            plan.bump("blank_po_rows")

    plan.bump("normalized_po_numbers", len({po for po in group_po.values() if po}))

    for source_id, group_rows in groups.items():
        po_normalized = group_po[source_id]
        quarantine_reasons: list[str] = []
        identity: dict[str, str | None] = {}
        for letter in HEADER_IDENTITY_COLUMNS:
            value, conflicted = _single_value(row.text(letter) for row in group_rows)
            if conflicted:
                quarantine_reasons.append(f"header_identity_conflict:{letter}")
            identity[letter] = value

        payload: dict[str, Any] = {
            "production_order_number": (
                normalize_text(group_rows[0].cell(COL_IMPORT_PO))
                or f"GOOGLE-ROW-{group_rows[0].sheet_row}"
            ),
            "status": identity[COL_PO_STATUS],
        }
        metadata: dict[str, Any] = {
            "order_list_source": {
                "spreadsheet_id": SOURCE_SPREADSHEET_ID,
                "tab": SOURCE_TAB_NAME,
                "gid": SOURCE_TAB_GID,
                "sheet_rows": sorted(r.sheet_row for r in group_rows),
            },
            "ordering_company": identity[COL_COMPANY],
            "order_vendor_name": identity[COL_ORDER_VENDOR],
            "vendor_id_raw": identity[COL_VENDOR_ID],
            "customer_name": identity[COL_CUSTOMER],
        }

        field_conflicts: list[str] = []
        for letter, column_name, kind in (
            (COL_SEAL_CONTAINER_DAY, "seal_container_date", "date"),
            (COL_SENT_PO_DATE, "sent_po_date", "date"),
            (COL_VENDOR_DELIVERY, "vendor_delivery_date", "date"),
            (COL_ETD, "etd", "date"),
            (COL_ETA, "eta", "date"),
            (COL_WAREHOUSE_DAY, "warehouse_date", "date"),
            (COL_BOOK_CONTAINER, "booking_state", "text"),
            (COL_CONTAINER_BOOKING, "container_booking_group", "text"),
            (COL_MBL, "mbl", "text"),
            (COL_CLOSE_TRACKING, "close_tracking", "boolean"),
        ):
            value, conflicted = _single_value(row.raw(letter) for row in group_rows)
            if conflicted:
                field_conflicts.append(letter)
                plan.bump("header_field_conflicts")
                payload[column_name] = None if kind != "boolean" else False
                metadata.setdefault("header_field_conflicts", []).append(letter)
                continue
            if kind == "date":
                typed, raw = parse_date(value)
                payload[column_name] = typed
                if raw is not None and typed is None:
                    metadata.setdefault("invalid_header_values", {})[letter] = raw
                    plan.bump("invalid_header_dates")
            elif kind == "boolean":
                typed, raw = parse_marker(value)
                payload[column_name] = bool(typed)
                if raw is not None and typed is None:
                    metadata.setdefault("invalid_header_values", {})[letter] = raw
            else:
                payload[column_name] = value

        order = PlannedOrder(
            source_id=source_id,
            po_normalized=po_normalized,
            payload=payload,
            metadata=metadata,
            sheet_rows=sorted(r.sheet_row for r in group_rows),
            quarantined=bool(quarantine_reasons),
            quarantine_reasons=quarantine_reasons,
        )
        if order.quarantined:
            plan.bump("quarantined_orders")
            plan.bump("quarantined_rows", len(group_rows))
        plan.orders.append(order)

    plan.bump("planned_orders", len(plan.orders))

    # ---- 2. Expand rows into canonical lines -----------------------------------------
    for row in populated:
        po_normalized = normalize_po_number(row.cell(COL_IMPORT_PO))
        order_sid = header_source_id(po_normalized, row.sheet_row)
        shape = classify_row_shape(row)
        plan.bump(f"rows_{shape}")

        if shape == ROW_SHAPE_NEITHER:
            plan.rejected_rows.append(
                RejectedRow(row.sheet_row, shape, "no_sku_and_no_assortment", {})
            )
            plan.bump("rejected_rows")
            continue

        if shape == ROW_SHAPE_BOTH:
            # Three such rows. They need human reconciliation before either shape is
            # trusted; creating a line from one shape would invent history.
            plan.rejected_rows.append(
                RejectedRow(row.sheet_row, shape, "both_shapes_pending_review", {})
            )
            plan.bump("rejected_rows")
            continue

        if shape == ROW_SHAPE_DIRECT:
            plan.lines.append(_plan_direct_line(row, order_sid, catalog, plan))
            continue

        split = split_assortment(row)
        if split.rejected:
            plan.rejected_rows.append(
                RejectedRow(
                    row.sheet_row,
                    shape,
                    f"assortment_structurally_invalid:{split.reason}",
                    {"component_counts": dict(split.counts)},
                )
            )
            plan.bump("rejected_rows")
            plan.bump("structurally_invalid_assortment_rows")
            continue

        for component in split.components:
            plan.lines.append(
                _plan_component_line(row, component, order_sid, catalog, plan)
            )
        plan.bump("assortment_components", len(split.components))

    plan.bump("planned_lines", len(plan.lines))
    return plan


def _line_common(row: SourceRow) -> tuple[dict[str, Any], dict[str, Any]]:
    """Line-scoped fields shared by direct rows and assortment components."""
    order_depth, order_depth_raw = parse_number(row.cell(COL_ORDER_DEPTH))
    case_pack, case_pack_raw = parse_number(row.cell(COL_CASE_PACK))
    cases, cases_raw = parse_number(row.cell(COL_CASES))
    start_ship, start_ship_raw = parse_date(row.cell(COL_START_SHIP))
    cancel, cancel_raw = parse_date(row.cell(COL_CANCEL_DATE))
    cargo, cargo_raw = parse_date(row.cell(COL_CARGO_FORECAST))
    sample_reorder, sample_reorder_raw = parse_marker(row.cell(COL_CONTRACTUAL_SAMPLE))

    payload: dict[str, Any] = {
        "order_person": row.text(COL_ORDER_PERSON),
        "order_type": row.text(COL_ORDER_TYPE),
        "customer_suffix": row.text(COL_SUFFIX),
        "customer_po_number": row.text(COL_CUSTOMER_PO),
        "assortment_id": row.text(COL_ASSORTMENT_ID),
        "order_depth_inches": order_depth,
        "case_pack": case_pack,
        "cases_reported": cases,
        "ship_to": row.text(COL_SHIP_TO),
        "start_ship_date": start_ship,
        "start_ship_raw": start_ship_raw,
        "cancel_date": cancel,
        "cancel_raw": cancel_raw,
        "cargo_forecast_date": cargo,
        "cargo_forecast_raw": cargo_raw,
        "contractual_sample_reorder": bool(sample_reorder),
    }

    invalid: dict[str, str] = {}
    for letter, typed, raw in (
        (COL_ORDER_DEPTH, order_depth, order_depth_raw),
        (COL_CASE_PACK, case_pack, case_pack_raw),
        (COL_CASES, cases, cases_raw),
    ):
        if raw is not None and typed is None:
            invalid[letter] = raw

    metadata: dict[str, Any] = {
        "order_list_source": {
            "spreadsheet_id": SOURCE_SPREADSHEET_ID,
            "tab": SOURCE_TAB_NAME,
            "gid": SOURCE_TAB_GID,
            "sheet_row": row.sheet_row,
        },
        # Columns F, G, Q, AP are preserved as raw evidence only. F/G hold dates and
        # "No data" rather than vendor identities; Q is "N/A" on all 8,416 populated
        # rows; AP duplicates derived PO-helper output.
        "raw_evidence": {
            COL_DEFAULT_VENDOR: row.raw(COL_DEFAULT_VENDOR),
            COL_SAMPLED_VENDOR: row.raw(COL_SAMPLED_VENDOR),
            COL_SAMPLE_DEPTH: row.raw(COL_SAMPLE_DEPTH),
            COL_FOR_WRITING_PO: row.raw(COL_FOR_WRITING_PO),
            COL_ORDER_TAB_LINE: row.raw(COL_ORDER_TAB_LINE),
        },
        "contractual_sample_reorder_raw": sample_reorder_raw,
    }
    if invalid:
        metadata["invalid_line_values"] = invalid
    return payload, metadata


def _apply_match(
    payload: dict[str, Any],
    metadata: dict[str, Any],
    match: MatchResult,
    plan: ImportPlan,
) -> None:
    payload["item_id"] = match.item_id
    payload["master_data_match_status"] = match.status
    metadata["master_data_resolution"] = match.resolution
    if match.pending_reason:
        metadata["item_link_pending_reason"] = match.pending_reason
    if match.candidate_count > 1:
        metadata["master_data_candidate_count"] = match.candidate_count
    plan.bump(f"match_{match.status}")
    plan.bump(f"resolution_{match.resolution}")


def _plan_direct_line(
    row: SourceRow, order_sid: str, catalog: MasterDataCatalog, plan: ImportPlan
) -> PlannedLine:
    payload, metadata = _line_common(row)
    sku = row.text(COL_STYLE)
    style_type = normalize_style_type(row.cell(COL_LICENSOR_OR_GENERIC))
    quantity, quantity_raw = parse_number(row.cell(COL_QUANTITY))

    payload.update(
        {
            "line_number": None,
            "sku": sku,
            "quantity_ordered": quantity,
            "source_style_type": style_type,
            "assortment_component_ordinal": None,
            "test_report": row.text(COL_TEST_REPORT),
            "professional_photos": row.text(COL_PROFESSIONAL_PHOTOS),
        }
    )
    metadata["order_list_snapshot"] = {
        "sku": sku,
        "style_type": style_type,
        "description": row.text(COL_DESCRIPTION),
        "license_status": row.text(COL_LICENSE_STATUS),
        "licensor_or_generic_raw": row.raw(COL_LICENSOR_OR_GENERIC),
        "quantity_raw": quantity_raw,
        # Preserve the raw AR ("Sub SKU") cell as evidence. On the 2026-08-11 workbook this
        # is a Style# mirror (see sub_sku_mirrors_style); recording it keeps the source
        # observable rather than discarding it just because it did not drive the shape.
        "sub_sku_raw": row.raw(COL_SUB_SKU),
        "sub_sku_is_style_mirror": sub_sku_mirrors_style(row),
    }
    _apply_match(payload, metadata, resolve_match(normalize_sku(sku), style_type, catalog), plan)
    return PlannedLine(
        source_id=line_source_id(row.sheet_row),
        order_source_id=order_sid,
        sheet_row=row.sheet_row,
        ordinal=None,
        payload=payload,
        metadata=metadata,
    )


def _plan_component_line(
    row: SourceRow,
    component: AssortmentComponent,
    order_sid: str,
    catalog: MasterDataCatalog,
    plan: ImportPlan,
) -> PlannedLine:
    payload, metadata = _line_common(row)
    parent_quantity, parent_quantity_raw = parse_number(row.cell(COL_QUANTITY))

    payload.update(
        {
            "line_number": None,
            "sku": component.sku,
            # Column AS is entirely blank. A component quantity is NEVER derived from
            # the parent assortment quantity — that would invent order history.
            "quantity_ordered": None,
            "source_style_type": component.style_type,
            "assortment_component_ordinal": component.ordinal,
            "test_report": component.test_report,
            "professional_photos": component.professional_photos,
        }
    )
    metadata["order_list_snapshot"] = {
        "sku": component.sku,
        "style_type": component.style_type,
        "description": component.description,
        "license_status": component.license_status,
        "licensor_or_generic_raw": row.raw(COL_LICENSOR_OR_GENERIC),
        "assortment_parent_quantity_raw": parent_quantity_raw,
        "assortment_parent_quantity": (
            str(parent_quantity) if parent_quantity is not None else None
        ),
        "component_quantity_source": "absent_never_guessed",
    }
    _apply_match(
        payload, metadata, resolve_match(normalize_sku(component.sku), component.style_type, catalog), plan
    )
    return PlannedLine(
        source_id=component_source_id(row.sheet_row, component.ordinal),
        order_source_id=order_sid,
        sheet_row=row.sheet_row,
        ordinal=component.ordinal,
        payload=payload,
        metadata=metadata,
    )


# =====================================================================================
# Database gateway
#
# The apply step talks ONLY to this interface, so the whole idempotency contract is
# provable against an in-memory fake with no database, no secrets and no network.
# =====================================================================================
class ImportGateway:
    """Minimal write surface. Every method is addressed by deterministic source ref."""

    def order_id_for_source(self, source_id: str) -> str | None:
        raise NotImplementedError

    def line_id_for_source(self, source_id: str) -> str | None:
        raise NotImplementedError

    def read_order(self, order_id: str) -> Mapping[str, Any]:
        raise NotImplementedError

    def read_line(self, line_id: str) -> Mapping[str, Any]:
        raise NotImplementedError

    def insert_order(self, payload: Mapping[str, Any], source_id: str) -> str:
        raise NotImplementedError

    def update_order(self, order_id: str, payload: Mapping[str, Any]) -> None:
        raise NotImplementedError

    def insert_line(
        self, order_id: str, payload: Mapping[str, Any], source_id: str
    ) -> str:
        raise NotImplementedError

    def update_line(self, line_id: str, payload: Mapping[str, Any]) -> None:
        raise NotImplementedError

    def begin_batch(self) -> None:
        raise NotImplementedError

    def commit_batch(self) -> None:
        raise NotImplementedError

    def rollback_batch(self) -> None:
        raise NotImplementedError


def canonical(value: Any) -> Any:
    """Comparison form. Two payloads are 'the same row' iff their canonical forms match.

    Without this, ``Decimal('5')`` vs ``Decimal('5.0')`` or ``date`` vs ``'2026-01-01'``
    would make an unchanged row look changed and a second run would rewrite everything.
    """
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, (datetime,)):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {key: canonical(value[key]) for key in sorted(value)}
    if isinstance(value, (list, tuple)):
        return [canonical(item) for item in value]
    return value


def payload_differs(desired: Mapping[str, Any], current: Mapping[str, Any]) -> list[str]:
    """Names of the fields that actually changed. Only fields the importer owns."""
    changed = []
    for key, value in desired.items():
        if canonical(value) != canonical(current.get(key)):
            changed.append(key)
    return sorted(changed)


@dataclass
class ApplyResult:
    orders_inserted: int = 0
    orders_updated: int = 0
    orders_unchanged: int = 0
    orders_drifted: int = 0
    orders_skipped_quarantined: int = 0
    lines_inserted: int = 0
    lines_updated: int = 0
    lines_unchanged: int = 0
    lines_drifted: int = 0
    lines_skipped_quarantined: int = 0
    drift_details: list[dict[str, Any]] = field(default_factory=list)

    @property
    def changed_rows(self) -> int:
        """Total business rows written. A second identical run must make this zero."""
        return (
            self.orders_inserted
            + self.orders_updated
            + self.lines_inserted
            + self.lines_updated
        )


def apply_plan(
    plan: ImportPlan,
    gateway: ImportGateway,
    *,
    batch_size: int = DEFAULT_BATCH_SIZE,
    replace_source: bool = False,
    dry_run: bool = False,
    allow_replace: bool = True,
    target_guard: Callable[[], None] | None = None,
    single_transaction: bool = False,
) -> ApplyResult:
    """Write the plan. Idempotent by construction.

    Every row is located by its deterministic Google source ref:

    * no ref  -> INSERT the row and its source ref;
    * ref + identical payload -> DO NOTHING (this is what makes run two a no-op);
    * ref + different payload -> DRIFT. Counted and reported, and rewritten only when
      ``--replace-source`` is passed. Silently overwriting is how an import destroys
      edits staff made in the app after the first load.

    ``allow_replace=False`` makes replacement structurally impossible for this call: it
    raises before the first batch opens rather than trusting a caller-side check. The
    production path always passes it, so ``--replace-source`` cannot reach production
    even if the CLI guard were removed.

    ``target_guard`` is invoked before EVERY batch. For a real write the production
    path points it at :func:`assert_server_identity_unchanged`, which asks the SERVER
    for its own identity. Re-reading ``supabase/.temp/project-ref`` there would prove
    nothing: the psycopg connection is opened once, and a later ``supabase link``
    cannot move an already-open socket. What can move underneath a live run is the
    server on the other end of it.

    ``single_transaction=True`` (always used for real writes) commits ONCE, at the very
    end. Per-batch commits were rejected: the orders loop finishes entirely before the
    lines loop starts, so an abort partway through the lines would leave all 3,212
    orders present with only a fraction of their 24,010 lines — four applications would
    then read orders whose line sets are silently short, which is far worse than no
    import at all. The batch structure is kept, but as the CADENCE of the drift check
    rather than as a commit boundary. Any exception rolls the whole import back.
    """
    if replace_source and not allow_replace:
        raise ImporterError(
            "replace_source is not permitted for this run (allow_replace=False). "
            "Refusing before any write. --replace-source has no production form: an "
            "import that can silently rewrite production rows is not an import, it is "
            "an unreviewed data migration."
        )
    result = ApplyResult()
    order_ids: dict[str, str] = {}

    quarantined = [order for order in plan.orders if order.quarantined]
    result.orders_skipped_quarantined = len(quarantined)
    blocked = {order.source_id for order in quarantined}
    result.lines_skipped_quarantined = sum(
        1 for line in plan.lines if line.order_source_id in blocked
    )

    orders = plan.writable_orders()
    for start in range(0, len(orders), batch_size):
        chunk = orders[start : start + batch_size]
        if not single_transaction:
            gateway.begin_batch()
        try:
            # Inside the try: a guard failure must roll the transaction back, not
            # escape past the rollback and leave it open.
            if target_guard is not None:
                target_guard()
            for order in chunk:
                full = dict(order.payload)
                full["metadata"] = order.metadata
                existing = gateway.order_id_for_source(order.source_id)
                if existing is None:
                    if dry_run:
                        result.orders_inserted += 1
                        continue
                    order_ids[order.source_id] = gateway.insert_order(
                        full, order.source_id
                    )
                    result.orders_inserted += 1
                    continue
                order_ids[order.source_id] = existing
                changed = payload_differs(full, gateway.read_order(existing))
                if not changed:
                    result.orders_unchanged += 1
                elif replace_source:
                    if not dry_run:
                        gateway.update_order(existing, full)
                    result.orders_updated += 1
                else:
                    result.orders_drifted += 1
                    result.drift_details.append(
                        {"kind": "order", "source_id": order.source_id, "fields": changed}
                    )
        except Exception:
            # In single-transaction mode this rolls back the ENTIRE import, not just
            # this chunk, which is exactly the intent.
            gateway.rollback_batch()
            raise
        if not single_transaction:
            gateway.commit_batch()

    lines = plan.writable_lines()
    for start in range(0, len(lines), batch_size):
        chunk = lines[start : start + batch_size]
        if not single_transaction:
            gateway.begin_batch()
        try:
            # Inside the try: a guard failure must roll the transaction back, not
            # escape past the rollback and leave it open.
            if target_guard is not None:
                target_guard()
            for line in chunk:
                full = dict(line.payload)
                full["metadata"] = line.metadata
                existing = gateway.line_id_for_source(line.source_id)
                if existing is None:
                    if dry_run:
                        result.lines_inserted += 1
                        continue
                    order_id = order_ids.get(line.order_source_id)
                    if order_id is None:
                        order_id = gateway.order_id_for_source(line.order_source_id)
                    if order_id is None:
                        raise ImporterError(
                            "Refusing to create an orphan line: no canonical order for "
                            f"source ref {line.order_source_id!r}."
                        )
                    gateway.insert_line(order_id, full, line.source_id)
                    result.lines_inserted += 1
                    continue
                changed = payload_differs(full, gateway.read_line(existing))
                if not changed:
                    result.lines_unchanged += 1
                elif replace_source:
                    if not dry_run:
                        gateway.update_line(existing, full)
                    result.lines_updated += 1
                else:
                    result.lines_drifted += 1
                    result.drift_details.append(
                        {"kind": "line", "source_id": line.source_id, "fields": changed}
                    )
        except Exception:
            # In single-transaction mode this rolls back the ENTIRE import, not just
            # this chunk, which is exactly the intent.
            gateway.rollback_batch()
            raise
        if not single_transaction:
            gateway.commit_batch()

    if single_transaction:
        # One commit for the whole import. Until this succeeds, nothing is durable.
        # A failure HERE is indeterminate, not a rollback: the server may have
        # committed and only the acknowledgement lost. It gets its own exception type
        # so the report cannot claim "no rows were written".
        try:
            gateway.commit_batch()
        except Exception as error:
            raise CommitOutcomeUnknown(
                "The final COMMIT did not return successfully, so the outcome of this "
                f"import is UNKNOWN ({error}). The rows may or may not be durable. Do "
                "NOT rerun blindly: query "
                "plm.production_order_source_ref for source_system "
                f"{SOURCE_SYSTEM!r} and establish the real state first."
            ) from error

    return result


# =====================================================================================
# In-memory gateway. Used by the tests AND by --dry-run, so the dry run exercises the
# real apply_plan code path rather than a parallel pretend one.
# =====================================================================================
class InMemoryGateway(ImportGateway):
    def __init__(self) -> None:
        self.orders: dict[str, dict[str, Any]] = {}
        self.lines: dict[str, dict[str, Any]] = {}
        self.order_refs: dict[str, str] = {}
        self.line_refs: dict[str, str] = {}
        #: Append-only log of every write. Its length is the idempotency proof.
        self.write_log: list[tuple[str, str]] = []
        self.batches_committed = 0
        self.batches_rolled_back = 0
        self._sequence = 0

    def _next_id(self, prefix: str) -> str:
        self._sequence += 1
        return f"{prefix}-{self._sequence:06d}"

    def order_id_for_source(self, source_id: str) -> str | None:
        return self.order_refs.get(source_id)

    def line_id_for_source(self, source_id: str) -> str | None:
        return self.line_refs.get(source_id)

    def read_order(self, order_id: str) -> Mapping[str, Any]:
        return self.orders[order_id]

    def read_line(self, line_id: str) -> Mapping[str, Any]:
        return self.lines[line_id]

    def insert_order(self, payload: Mapping[str, Any], source_id: str) -> str:
        order_id = self._next_id("order")
        self.orders[order_id] = dict(payload)
        self.order_refs[source_id] = order_id
        self.write_log.append(("insert_order", source_id))
        return order_id

    def update_order(self, order_id: str, payload: Mapping[str, Any]) -> None:
        self.orders[order_id].update(payload)
        self.write_log.append(("update_order", order_id))

    def insert_line(
        self, order_id: str, payload: Mapping[str, Any], source_id: str
    ) -> str:
        line_id = self._next_id("line")
        record = dict(payload)
        record["production_order_id"] = order_id
        self.lines[line_id] = record
        self.line_refs[source_id] = line_id
        self.write_log.append(("insert_line", source_id))
        return line_id

    def update_line(self, line_id: str, payload: Mapping[str, Any]) -> None:
        self.lines[line_id].update(payload)
        self.write_log.append(("update_line", line_id))

    def begin_batch(self) -> None:
        return None

    def commit_batch(self) -> None:
        self.batches_committed += 1

    def rollback_batch(self) -> None:
        self.batches_rolled_back += 1


# =====================================================================================
# Postgres gateway
# =====================================================================================
ORDER_COLUMNS = (
    "production_order_number",
    "status",
    "seal_container_date",
    "sent_po_date",
    "vendor_delivery_date",
    "etd",
    "eta",
    "warehouse_date",
    "booking_state",
    "container_booking_group",
    "mbl",
    "close_tracking",
    "metadata",
)

LINE_COLUMNS = (
    "sku",
    "quantity_ordered",
    "item_id",
    "line_number",
    "order_person",
    "order_type",
    "customer_suffix",
    "customer_po_number",
    "assortment_id",
    "assortment_component_ordinal",
    "order_depth_inches",
    "case_pack",
    "cases_reported",
    "ship_to",
    "start_ship_date",
    "start_ship_raw",
    "cancel_date",
    "cancel_raw",
    "cargo_forecast_date",
    "cargo_forecast_raw",
    "test_report",
    "professional_photos",
    "contractual_sample_reorder",
    "source_style_type",
    "master_data_match_status",
    "metadata",
)


class PostgresGateway(ImportGateway):  # pragma: no cover - needs a live database
    """Real writes. Never constructed by the test suite, by design."""

    def __init__(self, connection: Any) -> None:
        self._connection = connection

    def _fetch_one(self, sql: str, params: Sequence[Any]) -> Any:
        with self._connection.cursor() as cursor:
            cursor.execute(sql, params)
            row = cursor.fetchone()
        return row

    def order_id_for_source(self, source_id: str) -> str | None:
        row = self._fetch_one(
            "select production_order_id from plm.production_order_source_ref "
            "where source_system = %s and source_id = %s",
            (SOURCE_SYSTEM, source_id),
        )
        return None if row is None else str(row[0])

    def line_id_for_source(self, source_id: str) -> str | None:
        row = self._fetch_one(
            "select production_order_line_id from plm.production_order_line_source_ref "
            "where source_system = %s and source_id = %s",
            (SOURCE_SYSTEM, source_id),
        )
        return None if row is None else str(row[0])

    def _read(self, table: str, columns: Sequence[str], row_id: str) -> Mapping[str, Any]:
        select = ", ".join(columns)
        row = self._fetch_one(f"select {select} from {table} where id = %s", (row_id,))
        if row is None:
            raise ImporterError(f"{table} row {row_id} vanished mid-import.")
        record = dict(zip(columns, row))
        metadata = record.get("metadata")
        if isinstance(metadata, str):
            record["metadata"] = json.loads(metadata)
        return record

    def read_order(self, order_id: str) -> Mapping[str, Any]:
        return self._read("plm.production_order", ORDER_COLUMNS, order_id)

    def read_line(self, line_id: str) -> Mapping[str, Any]:
        return self._read("plm.production_order_line", LINE_COLUMNS, line_id)

    @staticmethod
    def _bind(payload: Mapping[str, Any], columns: Sequence[str]) -> list[Any]:
        values = []
        for column in columns:
            value = payload.get(column)
            values.append(json.dumps(canonical(value)) if column == "metadata" else value)
        return values

    def insert_order(self, payload: Mapping[str, Any], source_id: str) -> str:
        placeholders = ", ".join(["%s"] * len(ORDER_COLUMNS))
        row = self._fetch_one(
            f"insert into plm.production_order ({', '.join(ORDER_COLUMNS)}) "
            f"values ({placeholders}) returning id",
            self._bind(payload, ORDER_COLUMNS),
        )
        order_id = str(row[0])
        with self._connection.cursor() as cursor:
            cursor.execute(
                "insert into plm.production_order_source_ref "
                "(production_order_id, source_system, source_id, is_primary) "
                "values (%s, %s, %s, true) on conflict do nothing",
                (order_id, SOURCE_SYSTEM, source_id),
            )
        return order_id

    def update_order(self, order_id: str, payload: Mapping[str, Any]) -> None:
        assignments = ", ".join(f"{column} = %s" for column in ORDER_COLUMNS)
        with self._connection.cursor() as cursor:
            cursor.execute(
                f"update plm.production_order set {assignments}, updated_at = now() "
                "where id = %s",
                [*self._bind(payload, ORDER_COLUMNS), order_id],
            )

    def insert_line(
        self, order_id: str, payload: Mapping[str, Any], source_id: str
    ) -> str:
        columns = ("production_order_id", *LINE_COLUMNS)
        placeholders = ", ".join(["%s"] * len(columns))
        row = self._fetch_one(
            f"insert into plm.production_order_line ({', '.join(columns)}) "
            f"values ({placeholders}) returning id",
            [order_id, *self._bind(payload, LINE_COLUMNS)],
        )
        line_id = str(row[0])
        with self._connection.cursor() as cursor:
            cursor.execute(
                "insert into plm.production_order_line_source_ref "
                "(production_order_line_id, source_system, source_id, is_primary) "
                "values (%s, %s, %s, true) on conflict do nothing",
                (line_id, SOURCE_SYSTEM, source_id),
            )
        return line_id

    def update_line(self, line_id: str, payload: Mapping[str, Any]) -> None:
        assignments = ", ".join(f"{column} = %s" for column in LINE_COLUMNS)
        with self._connection.cursor() as cursor:
            cursor.execute(
                f"update plm.production_order_line set {assignments}, updated_at = now() "
                "where id = %s",
                [*self._bind(payload, LINE_COLUMNS), line_id],
            )

    def begin_batch(self) -> None:
        return None

    def commit_batch(self) -> None:
        self._connection.commit()

    def rollback_batch(self) -> None:
        self._connection.rollback()


# =====================================================================================
# Target proof and checksum gate
# =====================================================================================
def sha256_file(path: Path, chunk_size: int = 1 << 20) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def assert_source_checksum(path: Path, expected: str) -> str:
    """Refuse any workbook whose SHA-256 is not the approved one."""
    actual = sha256_file(path)
    if actual.lower() != expected.strip().lower():
        raise ImporterError(
            "Source checksum mismatch. Refusing to import.\n"
            f"  expected: {expected.strip().lower()}\n"
            f"  actual:   {actual}\n"
            "This is not a warning. A different workbook means a different contract; "
            "re-run the Phase 0 profile before importing it."
        )
    return actual


def assert_write_target(project_ref_file: Path, *, mode: str) -> str:
    """Prove the write target from disk. Called before the connection AND every batch.

    Re-reading the file on every call is the point: it is what turns a mid-run
    ``supabase link`` at another database into an abort instead of a partial import
    landing in the wrong project.
    """
    if mode not in MODE_REQUIRED_PROJECT_REF:
        raise ImporterError(
            "Refusing to write without an explicit target mode. Pass exactly one of "
            "--preview or --production (or --dry-run, which writes nothing)."
        )
    required = MODE_REQUIRED_PROJECT_REF[mode]
    if not project_ref_file.exists():
        raise ImporterError(
            f"Cannot prove the write target: {project_ref_file} does not exist. "
            f"Run `supabase link` against {mode} ({required}) first."
        )
    ref = project_ref_file.read_text(encoding="utf-8").strip()
    if ref == required:
        return ref
    other = (
        "PRODUCTION" if ref == PRODUCTION_PROJECT_REF
        else "PREVIEW" if ref == PREVIEW_PROJECT_REF
        else "an unknown project"
    )
    raise ImporterError(
        f"{project_ref_file} reads {ref!r} ({other}), but --{mode} requires "
        f"{required!r}. Refusing to write. Nothing has been written."
    )


def assert_preview_target(project_ref_file: Path, *, preview: bool) -> str:
    """Backwards-compatible preview-only wrapper around :func:`assert_write_target`."""
    return assert_write_target(
        project_ref_file, mode=MODE_PREVIEW if preview else MODE_DRY_RUN
    )


# =====================================================================================
# Server identity — the causal link between the proven ref and the actual writes
# =====================================================================================
@dataclass(frozen=True)
class ServerIdentity:
    """What the SERVER and the live socket say about themselves.

    ``project_ref`` comes from the parameters libpq actually used
    (``connection.info.host`` / ``.user``), not from the connection string as we parsed
    it. ``fingerprint`` is a live server-side tuple; it is re-read before every batch,
    so a pooler that re-points at a different backend, a failover, or a restart aborts
    the run instead of quietly finishing it somewhere else.
    """

    project_ref: str | None
    host: str | None
    port: int | None
    database: str | None
    user: str | None
    fingerprint: tuple[str, ...]

    def redacted(self) -> str:
        """Safe to print: host, database and ref only. Never the user or password."""
        return f"host={self.host} db={self.database} ref={self.project_ref}"


def project_ref_from_connection(host: str | None, user: str | None) -> str | None:
    """Recover the Supabase project ref from the live connection parameters."""
    for candidate in (host or "",):
        for pattern in HOST_PROJECT_REF_PATTERNS:
            found = pattern.match(candidate.strip())
            if found:
                return found.group(1).lower()
    found = POOLER_USER_REF_PATTERN.match((user or "").strip())
    if found:
        return found.group(1).lower()
    return None


def read_server_fingerprint(connection: Any) -> tuple[str, ...]:
    """A live server-side identity tuple. All three functions are public in Postgres."""
    with connection.cursor() as cursor:
        cursor.execute(
            "select current_database(), "
            "coalesce(inet_server_addr()::text, 'local'), "
            "pg_postmaster_start_time()::text"
        )
        row = cursor.fetchone()
    return tuple(str(value) for value in row)


def read_server_identity(connection: Any) -> ServerIdentity:
    info = connection.info
    host = getattr(info, "host", None)
    user = getattr(info, "user", None)
    port = getattr(info, "port", None)
    return ServerIdentity(
        project_ref=project_ref_from_connection(host, user),
        host=host,
        port=int(port) if port not in (None, "") else None,
        database=getattr(info, "dbname", None),
        user=user,
        fingerprint=read_server_fingerprint(connection),
    )


def assert_server_is_target(
    connection: Any, *, expected_ref: str, mode: str
) -> ServerIdentity:
    """Refuse unless the SERVER we are connected to is the proven target.

    This is the gate that makes the ref file mean anything. Without it,
    ``supabase/.temp/project-ref`` proves only the CLI's link state while
    ``PRODUCTION_DATABASE_URL`` could open a completely different database — and the
    reconciliation report would then name the wrong destination as evidence.
    """
    identity = read_server_identity(connection)
    if identity.project_ref is None:
        raise ImporterError(
            "Cannot prove which Supabase project this connection actually opened. "
            f"The live connection parameters ({identity.redacted()}) carry no "
            "recognisable project ref: a Supabase host is `db.<ref>.supabase.co` and a "
            "pooler user is `postgres.<ref>`. Refusing to write to a database whose "
            "identity cannot be established. Nothing has been written."
        )
    if identity.project_ref != expected_ref:
        raise ImporterError(
            "TARGET MISMATCH. The proven project ref and the database this connection "
            "actually opened are different databases.\n"
            f"  supabase/.temp/project-ref (and --{mode}) require: {expected_ref}\n"
            f"  the live connection actually opened:               "
            f"{identity.project_ref}\n"
            "The ref file records the Supabase CLI link state; the connection string "
            "decides where rows land. They disagree, so the connection string is "
            "stale or wrong. Refusing. Nothing has been written."
        )
    if mode == MODE_PRODUCTION and identity.port == TRANSACTION_POOLER_PORT:
        raise ImporterError(
            f"Refusing to import through the transaction-mode pooler (port "
            f"{TRANSACTION_POOLER_PORT}). It cannot hold the single transaction this "
            "import runs in, so a failure would leave orders committed without their "
            "lines. Use the direct or session connection. Nothing has been written."
        )
    return identity


def assert_server_identity_unchanged(connection: Any, baseline: ServerIdentity) -> None:
    """Re-assert the SERVER's own identity. Called before every batch.

    Re-reading the ref file would prove nothing here: the psycopg connection is opened
    once and a later ``supabase link`` cannot move an open socket. What CAN move
    underneath a live run is the server on the other end of it.
    """
    current = read_server_fingerprint(connection)
    if current != baseline.fingerprint:
        raise ImporterError(
            "TARGET DRIFT DETECTED MID-RUN. The database on the other end of this "
            "connection is no longer the one that was proven before the import "
            "started (its identity tuple changed: database name, server address or "
            "postmaster start time). Aborting. The transaction has been rolled back."
        )


def assert_no_existing_source_refs(connection: Any) -> int:
    """Production precondition: no ``google_order_list`` rows may exist yet.

    Required by the approval package, and it doubles as a content-based proof that
    this is not preview: preview already holds 3,212 imported orders, so a stale
    ``PRODUCTION_DATABASE_URL`` pointing there fails here even if everything else
    somehow passed.
    """
    with connection.cursor() as cursor:
        cursor.execute(
            "select count(*) from plm.production_order_source_ref "
            "where source_system = %s",
            [SOURCE_SYSTEM],
        )
        count = int(cursor.fetchone()[0])
    if count:
        raise ImporterError(
            f"Refusing: {count} {SOURCE_SYSTEM!r} order source references already "
            "exist in this database. A first production import expects zero. Either "
            "the import already ran, or this connection is not pointing at the "
            "database you think it is. Nothing has been written."
        )
    return count


def acquire_import_lock(connection: Any) -> None:
    """Refuse to run while another import holds the lock.

    ``order_id_for_source`` is read-then-insert at READ COMMITTED. Two overlapping
    runs — plausible when someone reruns after an aborted attempt — would both see
    "no row" and both insert. A session advisory lock removes the race outright
    instead of hiding it behind ``on conflict do nothing``, which would swallow the
    evidence that two runs overlapped.
    """
    with connection.cursor() as cursor:
        cursor.execute("select pg_try_advisory_lock(%s)", [IMPORT_ADVISORY_LOCK_KEY])
        acquired = bool(cursor.fetchone()[0])
    if not acquired:
        raise ImporterError(
            "Another OrderList import holds the advisory lock on this database. "
            "Refusing to run two importers concurrently. Nothing has been written."
        )


def assert_reviewed_commit_is_checked_out(commit_sha: str, repo_root: Path) -> None:
    """Prove the code being RUN is the code that was reviewed.

    Without this, ``--reviewed-commit`` is an unchecked operator string and, since the
    other two components of the confirmation are compile-time constants, the whole
    confirmation is self-issued and a stale approval is replayable. Fails closed: if
    git cannot answer, the run is refused.
    """
    import subprocess

    def git(*args: str) -> str:
        try:
            done = subprocess.run(
                ["git", "-C", str(repo_root), *args],
                capture_output=True,
                text=True,
                check=False,
            )
        except OSError as error:
            raise ImporterError(
                f"Cannot verify --reviewed-commit: git is not runnable ({error}). "
                "Refusing. Nothing has been written."
            ) from error
        if done.returncode != 0:
            raise ImporterError(
                f"Cannot verify --reviewed-commit: `git {' '.join(args)}` failed "
                f"({done.stderr.strip()}). Refusing. Nothing has been written."
            )
        return done.stdout.strip()

    head = git("rev-parse", "HEAD").lower()
    if head != commit_sha:
        raise ImporterError(
            "--reviewed-commit does not match the checked-out code.\n"
            f"  --reviewed-commit: {commit_sha}\n"
            f"  HEAD:              {head}\n"
            "A confirmation is only meaningful if the importer being run IS the "
            "reviewed importer. Check out the reviewed commit. Nothing has been "
            "written."
        )
    # --untracked-files=no is deliberate and load-bearing. Plain --porcelain lists
    # untracked files as "??", and a normal checkout always has some — the live
    # orchestrator checkout has 18 today. Worse, this gate would be self-inflicting:
    # an aborted run writes its abort report to an UNTRACKED path, so the next attempt
    # would be refused because of the evidence the importer itself just produced, and
    # the only way to proceed would be to delete that evidence or commit mid-incident.
    # A gate whose workaround is "destroy the abort report" is worse than no gate.
    # Every modification to reviewed, tracked code is still caught.
    dirty = git("status", "--porcelain", "--untracked-files=no")
    if dirty:
        raise ImporterError(
            "Tracked files are modified, so the code being run is not exactly the "
            f"reviewed commit {commit_sha}. Refusing. Nothing has been written."
        )


# =====================================================================================
# Production confirmation gate
# =====================================================================================
def build_production_confirmation(
    *, commit_sha: str, source_sha256: str, project_ref: str
) -> str:
    """The exact sentence a production run must be given via ``--confirm``.

    Deterministic and lower-cased in its variable parts so the operator can regenerate
    it, but NOT derivable from the flags alone at check time: the values it is built
    from are the ones actually proven at runtime (the hash of the file on disk, the ref
    read from ``supabase/.temp/project-ref``), never the ones the operator typed.
    """
    return PRODUCTION_CONFIRMATION_TEMPLATE.format(
        source_sha256=source_sha256.strip().lower(),
        project_ref=project_ref.strip(),
        commit_sha=commit_sha.strip().lower(),
    )


def _normalize_confirmation(value: str) -> str:
    """Collapse whitespace, case and a trailing period.

    A correct approval pasted out of a wrapped document, or ended with a full stop,
    must not be rejected — otherwise the operator learns to retype it, which is exactly
    how a wrong confirmation gets typed with confidence. Nothing SEMANTIC is relaxed:
    the commit, the hash and the ref must still match character for character.
    """
    collapsed = " ".join(value.split()).strip()
    while collapsed.endswith(".") or collapsed.endswith(" "):
        collapsed = collapsed[:-1].rstrip()
    return collapsed.lower()


def assert_reviewed_commit(commit_sha: str | None) -> str:
    """A production run must name the full 40-character reviewed commit."""
    candidate = (commit_sha or "").strip().lower()
    if not candidate:
        raise ImporterError(
            "--reviewed-commit is required for --production. It must be the full "
            "40-character shared-db commit SHA that was reviewed and merged, and it "
            "must be the commit this working tree is checked out at."
        )
    if len(candidate) != FULL_GIT_SHA_LENGTH or not re.fullmatch(
        r"[0-9a-f]{40}", candidate
    ):
        raise ImporterError(
            f"--reviewed-commit {commit_sha!r} is not a full 40-character git SHA. "
            "An abbreviated SHA is not an exact approval package. Refusing."
        )
    return candidate


def assert_production_confirmation(
    provided: str | None, *, commit_sha: str, source_sha256: str, project_ref: str
) -> str:
    """Refuse a production run whose confirmation is missing or does not match.

    Raises BEFORE any credential is read and before any connection is opened.
    """
    expected = build_production_confirmation(
        commit_sha=commit_sha, source_sha256=source_sha256, project_ref=project_ref
    )
    if not (provided or "").strip():
        raise ImporterError(
            "--confirm is required for --production and was empty.\n"
            "The exact confirmation for THIS commit, workbook and target is:\n"
            f"  {expected}\n"
            "Nothing has been written."
        )
    if _normalize_confirmation(provided or "") != _normalize_confirmation(expected):
        raise ImporterError(
            "--confirm does not match this run. Refusing before any credential is "
            "used.\n"
            f"  expected: {expected}\n"
            "  received: (withheld — it did not match, and echoing it back would "
            "invite copy-pasting the wrong approval into the next attempt)\n"
            "A confirmation is bound to one reviewed commit, one workbook hash and "
            "one project ref. Nothing has been written."
        )
    return expected


# =====================================================================================
# Reconciliation report
# =====================================================================================
@dataclass
class Reconciliation:
    counters: Mapping[str, int]
    apply_result: ApplyResult
    baseline: ReconciliationBaseline = BASELINE
    checksum_matched_approved_source: bool = False
    #: Operator-declared populated-row count. Required for production. It is a CHECKED
    #: INPUT, never a value the code picks for itself. The approved figure is 12,354
    #: (owner ruling; the question is closed). Requiring it on the command line keeps
    #: the approved number visible in the approved command and asserted before any
    #: write, instead of being inferred from a constant.
    expected_populated_rows: int | None = None

    def balance_checks(self) -> list[tuple[str, bool, str]]:
        """``(name, passed, detail)``. These are ASSERTIONS, not decoration."""
        counters = self.counters
        staged = counters.get("staged_rows", 0)
        shapes = sum(
            counters.get(f"rows_{shape}", 0)
            for shape in (
                ROW_SHAPE_DIRECT,
                ROW_SHAPE_ASSORTMENT,
                ROW_SHAPE_BOTH,
                ROW_SHAPE_NEITHER,
            )
        )
        matches = sum(
            counters.get(f"match_{status}", 0)
            for status in (
                MATCH_MATCHED,
                MATCH_UNMATCHED,
                MATCH_AMBIGUOUS,
                MATCH_MANUAL,
                MATCH_NOT_APPLICABLE,
            )
        )
        planned_lines = counters.get("planned_lines", 0)

        checks: list[tuple[str, bool, str]] = [
            (
                "staged rows equal the sum of the four row shapes",
                staged == shapes,
                f"{staged} staged vs {shapes} classified",
            ),
            (
                "every planned line carries exactly one match status",
                matches == planned_lines,
                f"{matches} statuses vs {planned_lines} lines",
            ),
        ]

        if self.expected_populated_rows is not None:
            checks.append(
                (
                    "staged rows equal the operator-declared expected populated rows",
                    staged == self.expected_populated_rows,
                    f"{staged} staged vs {self.expected_populated_rows} declared",
                )
            )

        if self.checksum_matched_approved_source:
            base = self.baseline
            checks.extend(
                [
                    (
                        "staged rows equal the Phase 0 populated-row baseline",
                        staged == base.populated_rows,
                        f"{staged} vs {base.populated_rows}",
                    ),
                    (
                        "direct-only rows equal the Phase 0 baseline",
                        counters.get(f"rows_{ROW_SHAPE_DIRECT}", 0)
                        == base.direct_only_rows,
                        f"{counters.get(f'rows_{ROW_SHAPE_DIRECT}', 0)} vs "
                        f"{base.direct_only_rows}",
                    ),
                    (
                        "assortment-only rows equal the Phase 0 baseline",
                        counters.get(f"rows_{ROW_SHAPE_ASSORTMENT}", 0)
                        == base.assortment_only_rows,
                        f"{counters.get(f'rows_{ROW_SHAPE_ASSORTMENT}', 0)} vs "
                        f"{base.assortment_only_rows}",
                    ),
                    (
                        "both-shape rows equal the Phase 0 baseline",
                        counters.get(f"rows_{ROW_SHAPE_BOTH}", 0) == base.both_shape_rows,
                        f"{counters.get(f'rows_{ROW_SHAPE_BOTH}', 0)} vs "
                        f"{base.both_shape_rows}",
                    ),
                    (
                        "neither-shape rows equal the Phase 0 baseline",
                        counters.get(f"rows_{ROW_SHAPE_NEITHER}", 0)
                        == base.neither_shape_rows,
                        f"{counters.get(f'rows_{ROW_SHAPE_NEITHER}', 0)} vs "
                        f"{base.neither_shape_rows}",
                    ),
                    (
                        "assortment components equal the Phase 0 baseline",
                        counters.get("assortment_components", 0)
                        == base.assortment_components,
                        f"{counters.get('assortment_components', 0)} vs "
                        f"{base.assortment_components}",
                    ),
                    (
                        "blank-PO rows equal the Phase 0 baseline",
                        counters.get("blank_po_rows", 0) == base.blank_po_rows,
                        f"{counters.get('blank_po_rows', 0)} vs {base.blank_po_rows}",
                    ),
                    (
                        "normalized PO numbers equal the Phase 0 baseline",
                        counters.get("normalized_po_numbers", 0)
                        == base.normalized_po_numbers,
                        f"{counters.get('normalized_po_numbers', 0)} vs "
                        f"{base.normalized_po_numbers}",
                    ),
                ]
            )
        return checks

    def balanced(self) -> bool:
        return all(passed for _, passed, _ in self.balance_checks())


def render_report(
    *,
    reconciliation: Reconciliation,
    checksum: str,
    project_ref: str,
    workbook_name: str,
    run_mode: str,
    generated_at: str,
    second_run: ApplyResult | None = None,
    title: str = "PopDAM OrderList preview import reconciliation",
    reviewed_commit: str | None = None,
) -> str:
    """Markdown reconciliation report.

    Secret-free and customer-free by construction: it emits counts and deterministic
    source refs only. No SKU, customer, vendor, PO text or credential ever reaches it.
    """
    counters = reconciliation.counters
    applied = reconciliation.apply_result
    lines: list[str] = []
    add = lines.append

    add(f"# {title}")
    add("")
    add(f"- Generated: {generated_at}")
    add(f"- Run mode: **{run_mode}**")
    if reviewed_commit:
        add(f"- Reviewed shared-db commit: `{reviewed_commit}`")
    # NB: the label deliberately avoids the word "Target", which is also a customer
    # name the leak test in scripts/tests/test_import_order_list.py screens for.
    add(f"- Destination project ref: `{project_ref}`")
    add(f"- Source spreadsheet: `{SOURCE_SPREADSHEET_ID}`, tab `{SOURCE_TAB_NAME}` "
        f"(gid `{SOURCE_TAB_GID}`)")
    add(f"- Workbook file: `{workbook_name}`")
    add(f"- Workbook SHA-256: `{checksum}`")
    add(
        "- Matches approved Phase 0 export: "
        f"**{'yes' if reconciliation.checksum_matched_approved_source else 'NO'}**"
    )
    add(f"- Source system: `{SOURCE_SYSTEM}`")
    add("")
    add("This report contains counts and deterministic source references only. "
        "No customer, vendor, SKU or order text is reproduced here.")
    add("")

    add("## Row shapes")
    add("")
    add("| Class | Rows |")
    add("|---|---:|")
    for label, key in (
        ("Staged (populated) rows", "staged_rows"),
        ("Direct SKU rows", f"rows_{ROW_SHAPE_DIRECT}"),
        ("Assortment rows", f"rows_{ROW_SHAPE_ASSORTMENT}"),
        ("Both shapes (rejected, needs review)", f"rows_{ROW_SHAPE_BOTH}"),
        ("Neither shape (rejected)", f"rows_{ROW_SHAPE_NEITHER}"),
        ("Structurally invalid assortment rows", "structurally_invalid_assortment_rows"),
        ("Assortment components expanded", "assortment_components"),
        ("Rejected rows total", "rejected_rows"),
    ):
        add(f"| {label} | {counters.get(key, 0)} |")
    add("")

    add("## Orders and headers")
    add("")
    add("| Measure | Count |")
    add("|---|---:|")
    for label, key in (
        ("Planned canonical orders", "planned_orders"),
        ("Normalized PO numbers", "normalized_po_numbers"),
        ("Blank-PO rows (each its own header)", "blank_po_rows"),
        ("Quarantined orders (identity conflict)", "quarantined_orders"),
        ("Rows inside quarantined orders", "quarantined_rows"),
        ("Header fields left NULL on disagreement", "header_field_conflicts"),
        ("Invalid header dates preserved as raw evidence", "invalid_header_dates"),
    ):
        add(f"| {label} | {counters.get(key, 0)} |")
    add("")

    add("## Master Data resolution")
    add("")
    add("`master_data_match_status` is the database column. `resolution` is the Master "
        "Data evidence, which survives even while `plm.item` is unpopulated.")
    add("")
    add("| Measure | Count |")
    add("|---|---:|")
    for label, key in (
        ("Planned canonical lines", "planned_lines"),
        ("status = matched", f"match_{MATCH_MATCHED}"),
        ("status = unmatched", f"match_{MATCH_UNMATCHED}"),
        ("status = ambiguous", f"match_{MATCH_AMBIGUOUS}"),
        ("status = not_applicable", f"match_{MATCH_NOT_APPLICABLE}"),
        ("resolution = unique Master Data row", f"resolution_{RESOLUTION_UNIQUE}"),
        ("resolution = ambiguous Master Data rows", f"resolution_{RESOLUTION_AMBIGUOUS}"),
        ("resolution = unmatched SKU", f"resolution_{RESOLUTION_UNMATCHED}"),
        ("resolution = no usable SKU/type", f"resolution_{RESOLUTION_NO_TYPE}"),
    ):
        add(f"| {label} | {counters.get(key, 0)} |")
    add("")
    add("> `item_id` is NULL wherever `resolution = unique` but `plm.item` is still "
        "empty. That is issue #727's coordination constraint, not a defect: "
        "fix_schema_for_api.md Phase 4 owns populating `plm.item`.")
    add("")

    add("## Writes")
    add("")
    add("| Measure | First run |" + (" Second run |" if second_run else ""))
    add("|---|---:|" + ("---:|" if second_run else ""))
    for label, attribute in (
        ("Orders inserted", "orders_inserted"),
        ("Orders updated", "orders_updated"),
        ("Orders unchanged", "orders_unchanged"),
        ("Orders drifted (not rewritten)", "orders_drifted"),
        ("Orders skipped (quarantined)", "orders_skipped_quarantined"),
        ("Lines inserted", "lines_inserted"),
        ("Lines updated", "lines_updated"),
        ("Lines unchanged", "lines_unchanged"),
        ("Lines drifted (not rewritten)", "lines_drifted"),
        ("Lines skipped (quarantined)", "lines_skipped_quarantined"),
    ):
        row = f"| {label} | {getattr(applied, attribute)} |"
        if second_run:
            row += f" {getattr(second_run, attribute)} |"
        add(row)
    add("")
    if second_run:
        add(
            f"**Idempotency: the second identical run changed "
            f"{second_run.changed_rows} business rows.** "
            + ("PASS." if second_run.changed_rows == 0 else "FAIL — investigate.")
        )
        add("")

    add("## Balance checks")
    add("")
    add("| Check | Result | Detail |")
    add("|---|---|---|")
    for name, passed, detail in reconciliation.balance_checks():
        add(f"| {name} | {'PASS' if passed else 'FAIL'} | {detail} |")
    add("")
    add(f"**Overall: {'BALANCED' if reconciliation.balanced() else 'NOT BALANCED'}**")
    add("")

    add("## Source row-count history (provenance; settled by owner ruling)")
    add("")
    add(
        "The approved populated-row count is **12,354**. Two earlier figures appear in "
        "the history and are recorded here as provenance only — the owner closed the "
        "question by decision and no itemisation of the difference is required. This "
        "run still asserts against the count stated on the command line."
    )
    add("")
    add("| Count | Workbook SHA-256 | How it was produced |")
    add("|---:|---|---|")
    add(
        "| 12,328 | `4958b4b7…fc6abbe4` | 2026-08-09 Phase 0 audit of the "
        "then-approved export. |"
    )
    add(
        "| 12,323 | `b9b282dc…ffba30` / `904b2cb9…0b2845` | 2026-08-11 browser exports "
        "of the live sheet. DIFFERENT files from both the 2026-08-09 export and the "
        "approved one; neither was retained. |"
    )
    add(
        f"| 12,354 | `{APPROVED_SOURCE_SHA256}` | 2026-08-12 re-profile of the "
        "owner-accepted workbook using this importer's own `is_populated` definition "
        "(any mapped cell non-empty). The same file gives 12,349 under a PO-Status "
        "definition. |"
    )
    add("")
    add(
        "For the record: the five-row shrink (12,328 -> 12,323) was measured on a file "
        "that is neither the 2026-08-09 export nor the approved workbook, and that file "
        "was not kept, so which rows changed cannot be recovered from this repository. "
        "The owner closed this by ruling that 12,354 is the approved count. It is not "
        "an outstanding gate."
    )
    add("")

    if applied.drift_details:
        add("## Drift (existing rows whose source payload changed)")
        add("")
        add("These rows were NOT rewritten. Re-run with `--replace-source` (preview "
            "only) after reviewing them.")
        add("")
        add("| Kind | Source ref | Changed fields |")
        add("|---|---|---|")
        for detail in applied.drift_details[:200]:
            fields = ", ".join(detail["fields"])
            add(f"| {detail['kind']} | `{detail['source_id']}` | {fields} |")
        if len(applied.drift_details) > 200:
            add(f"| ... | ... | {len(applied.drift_details) - 200} more omitted |")
        add("")

    return "\n".join(lines) + "\n"


# =====================================================================================
# Workbook reading
# =====================================================================================
def read_workbook_rows(
    path: Path, sheet_name: str = SOURCE_TAB_NAME, *, header_rows: int = 1
) -> list[SourceRow]:
    """Read the ``Order`` tab. Values only, read-only, never writes the file back."""
    try:
        from openpyxl import load_workbook  # type: ignore import-not-found
    except ImportError as error:  # pragma: no cover - environment dependent
        raise ImporterError(
            "openpyxl is required to read the workbook. Install it with "
            "`pip install openpyxl`. It is NOT required to run the test suite."
        ) from error

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ImporterError(
                f"Workbook has no {sheet_name!r} tab. Found: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
        width = column_index(LAST_COLUMN) + 1
        rows: list[SourceRow] = []
        for number, values in enumerate(
            sheet.iter_rows(values_only=True, max_col=width), start=1
        ):
            if number <= header_rows:
                continue
            padded = tuple(values) + (None,) * (width - len(values))
            rows.append(SourceRow(sheet_row=number, values=padded[:width]))
        return rows
    finally:
        workbook.close()


# =====================================================================================
# CLI
# =====================================================================================
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Idempotent importer for the legacy Google OrderList Order tab. "
            "Writes only with an explicit --preview or --production flag; "
            "--production additionally requires the approved workbook hash, the "
            "reviewed commit, and an exact confirmation."
        )
    )
    parser.add_argument("--workbook", required=True, type=Path, help="Local .xlsx export")
    parser.add_argument(
        "--expected-sha256",
        default=os.environ.get("ORDER_LIST_SHA256", APPROVED_SOURCE_SHA256),
        help=(
            "Required source checksum. Defaults to the approved Phase 0 export "
            f"({APPROVED_SOURCE_SHA256})."
        ),
    )
    parser.add_argument("--sheet", default=SOURCE_TAB_NAME)
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Plan and reconcile against an in-memory database. Writes nothing.",
    )
    target = parser.add_mutually_exclusive_group()
    target.add_argument(
        "--preview",
        action="store_true",
        help=f"Write to preview. Target must be {PREVIEW_PROJECT_REF}.",
    )
    target.add_argument(
        "--production",
        action="store_true",
        help=(
            f"Write to PRODUCTION ({PRODUCTION_PROJECT_REF}). Requires the approved "
            "workbook hash, --reviewed-commit, --confirm and "
            "--expected-populated-rows. Cannot be combined with --preview or "
            "--replace-source."
        ),
    )
    parser.add_argument(
        "--reviewed-commit",
        default=None,
        help=(
            "Full 40-character shared-db commit SHA that was reviewed and merged. "
            "Required for --production."
        ),
    )
    parser.add_argument(
        "--confirm",
        default=None,
        help=(
            "Exact confirmation sentence bound to the reviewed commit, the workbook "
            "SHA-256 and the project ref. Required for --production. A wrong or "
            "missing value aborts before the connection string is read."
        ),
    )
    parser.add_argument(
        "--expected-populated-rows",
        type=int,
        default=None,
        help=(
            "The populated-row count the operator expects. Required for --production "
            "and asserted as a balance check. Deliberately NOT defaulted, so the "
            "approved figure (12,354) is visible in the approved command itself."
        ),
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help=(
            "Connection string. Never logged. Defaults to PREVIEW_DATABASE_URL for "
            "--preview and PRODUCTION_DATABASE_URL for --production; it is read only "
            "AFTER every gate has passed."
        ),
    )
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument(
        "--project-ref-file", type=Path, default=Path(DEFAULT_PROJECT_REF_FILE)
    )
    parser.add_argument(
        "--replace-source",
        action="store_true",
        help=(
            "Rewrite existing rows whose source payload drifted. PREVIEW ONLY, and "
            "only after reviewing the drift table from a run without it. There is no "
            "production form of this flag."
        ),
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=None,
        help="Defaults to docs/verification/popdam-order-list-preview-<YYYY-MM-DD>/",
    )
    parser.add_argument(
        "--verify-idempotency",
        action="store_true",
        help="Run apply twice and record that the second run changed zero rows.",
    )
    return parser


def load_master_data_catalog(connection: Any) -> MasterDataCatalog:  # pragma: no cover
    """Build the catalog from Master Data plus the style-to-item bridge.

    ``plm.item`` is empty today, so ``item_ids`` comes back all-NULL and every uniquely
    resolved line is recorded as ``unmatched`` with
    ``metadata.item_link_pending_reason = 'plm_item_unpopulated'``. Do NOT "fix" that by
    inventing items.
    """
    rows: dict[tuple[str, str], list[str]] = {}
    item_ids: dict[str, str | None] = {}
    with connection.cursor() as cursor:
        cursor.execute(
            """
            select
              str.id::text,
              nullif(btrim(lower(str.sku)), ''),
              case str.source_sheet
                when 'License.Style' then 'licensed'
                when 'Generic.Style' then 'generic'
              end,
              bridge.plm_item_id::text
            from public.style_tracker_rows str
            left join plm.style_tracker_item_bridge bridge
              on bridge.style_tracker_row_id = str.id
            where str.source_sheet in ('License.Style', 'Generic.Style')
            """
        )
        for row_id, sku, style_type, item_id in cursor.fetchall():
            if not sku or not style_type:
                continue
            rows.setdefault((sku, style_type), []).append(row_id)
            item_ids[row_id] = item_id
    return MasterDataCatalog(rows=rows, item_ids=item_ids)


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    # -------------------------------------------------------------------------------
    # Gate order matters. Everything that can refuse this run is checked BEFORE the
    # connection string is read, so a wrong invocation never touches a credential.
    # -------------------------------------------------------------------------------
    if args.preview and args.production:  # pragma: no cover - argparse blocks it first
        raise ImporterError("--preview and --production are mutually exclusive.")
    if args.dry_run and (args.preview or args.production):
        # A "--dry-run --production" run would open a real production connection, run
        # live queries, and then write a report titled "production import
        # reconciliation" with run mode "production write" claiming 3,212 orders and
        # 24,010 lines inserted — for an import that never happened — over the top of
        # any genuine same-day report. In a repo whose approval process runs on these
        # documents that is a manufactured false record. Refused.
        raise ImporterError(
            "--dry-run cannot be combined with --preview or --production. A dry run is "
            "in-memory and opens no connection; combining them would produce a report "
            "that claims an import which never happened. Run --dry-run on its own for "
            "the rehearsal: every asserted Phase 0 baseline count is derived from the "
            "workbook, not from the live catalog, so the rehearsal loses nothing it "
            "is required to prove."
        )
    if not args.dry_run and not (args.preview or args.production):
        raise ImporterError(
            "Pass --dry-run, --preview or --production. Refusing to guess."
        )
    if args.replace_source and args.production:
        raise ImporterError(
            "--replace-source is refused with --production. There is no production "
            "form of replacement: rewriting production rows from a spreadsheet is an "
            "unreviewed data migration, not an import. Nothing has been written."
        )
    if args.replace_source and not args.preview:
        raise ImporterError("--replace-source is preview-only.")
    if args.batch_size < 1:
        raise ImporterError("--batch-size must be at least 1.")

    mode = (
        MODE_PRODUCTION if args.production
        else MODE_PREVIEW if args.preview
        else MODE_DRY_RUN
    )

    reviewed_commit: str | None = None
    expected_sha256 = args.expected_sha256
    expected_populated_rows = args.expected_populated_rows
    if mode == MODE_PRODUCTION:
        # The hash is PINNED, not merely defaulted. --expected-sha256 (and the
        # ORDER_LIST_SHA256 environment variable behind it) cannot relax production.
        if expected_sha256.strip().lower() != APPROVED_SOURCE_SHA256:
            raise ImporterError(
                "--production pins the workbook SHA-256 to the approved constant "
                f"{APPROVED_SOURCE_SHA256}. It cannot be overridden by "
                "--expected-sha256 or ORDER_LIST_SHA256. Nothing has been written."
            )
        expected_sha256 = APPROVED_SOURCE_SHA256
        reviewed_commit = assert_reviewed_commit(args.reviewed_commit)
        assert_reviewed_commit_is_checked_out(
            reviewed_commit, Path(__file__).resolve().parents[1]
        )
        if args.project_ref_file != Path(DEFAULT_PROJECT_REF_FILE):
            # Historically the whole target proof reduced to "a file containing the
            # right string", which an operator could point anywhere. The server-identity
            # gate below is now the real proof, but there is no reason to keep this
            # override available in production either.
            raise ImporterError(
                "--project-ref-file cannot be overridden with --production. The "
                f"default {DEFAULT_PROJECT_REF_FILE} is the only accepted location."
            )
        if expected_populated_rows is None:
            raise ImporterError(
                "--expected-populated-rows is required for --production. The source "
                "approved figure is 12,354; state it explicitly so the command being "
                "approved carries the number it is allowed to see."
            )
    if expected_populated_rows is not None and expected_populated_rows < 1:
        raise ImporterError("--expected-populated-rows must be at least 1.")

    workbook = args.workbook
    if not workbook.exists():
        raise ImporterError(f"Workbook not found: {workbook}")

    checksum = assert_source_checksum(workbook, expected_sha256)
    matches_approved = checksum.lower() == APPROVED_SOURCE_SHA256
    print(f"Source checksum verified: {checksum}", file=sys.stderr)

    # Prove the target from disk BEFORE the confirmation, so the confirmation is
    # checked against the ref actually on disk rather than the one the operator hoped
    # for. Re-proved again below, immediately before the connection opens, and once
    # more before every batch.
    project_ref: str
    target_guard: Callable[[], None] | None = None
    if mode in MODE_REQUIRED_PROJECT_REF:
        project_ref = assert_write_target(args.project_ref_file, mode=mode)
    else:
        project_ref = "(none — dry run)"

    if mode == MODE_PRODUCTION:
        assert reviewed_commit is not None
        assert_production_confirmation(
            args.confirm,
            commit_sha=reviewed_commit,
            source_sha256=checksum,
            project_ref=project_ref,
        )
        print("Production confirmation accepted.", file=sys.stderr)

    rows = read_workbook_rows(workbook, args.sheet)
    print(f"Read {len(rows)} physical rows from tab {args.sheet!r}.", file=sys.stderr)

    if mode in MODE_REQUIRED_PROJECT_REF:
        # Recheck immediately before the transaction opens. Between the gates above and
        # here the workbook was read, which takes real time; a relink in that window
        # must abort, not be inherited.
        project_ref = assert_write_target(args.project_ref_file, mode=mode)
        print(f"Link state proven: {project_ref}", file=sys.stderr)

        env_var = (
            "PRODUCTION_DATABASE_URL" if mode == MODE_PRODUCTION
            else "PREVIEW_DATABASE_URL"
        )
        database_url = args.database_url or os.environ.get(env_var)
        if not database_url:
            raise ImporterError(f"--database-url (or {env_var}) is required.")
        try:
            import psycopg  # type: ignore import-not-found
        except ImportError as error:  # pragma: no cover
            raise ImporterError(
                f"psycopg is required for --{mode}. `pip install 'psycopg[binary]'`."
            ) from error
        connection = psycopg.connect(database_url, autocommit=False)

        # THE causal gate. Everything above proves the Supabase CLI's link state and
        # the operator's intent; only this proves that the database on the other end of
        # THIS socket is the one the report is about to name as the destination.
        identity = assert_server_is_target(
            connection, expected_ref=project_ref, mode=mode
        )
        print(f"Server identity proven: {identity.redacted()}", file=sys.stderr)
        acquire_import_lock(connection)
        if mode == MODE_PRODUCTION:
            assert_no_existing_source_refs(connection)

        def target_guard() -> None:
            assert_server_identity_unchanged(connection, identity)

        catalog = load_master_data_catalog(connection)
        gateway: ImportGateway = PostgresGateway(connection)
        run_mode = f"{mode} write"
    else:
        catalog = MasterDataCatalog()
        gateway = InMemoryGateway()
        run_mode = "dry run (in-memory, nothing written)"

    allow_replace = mode != MODE_PRODUCTION
    writing = mode in MODE_REQUIRED_PROJECT_REF

    plan = build_plan(rows, catalog)

    # -------------------------------------------------------------------------------
    # THE BALANCE CHECKS GATE THE WRITE. They used to run after every batch was
    # committed, which meant an approver who passed the wrong --expected-populated-rows
    # got all 3,212 orders and 24,010 lines inserted and only then read "did NOT
    # balance", with nothing rolled back. Every counter these checks assert on comes
    # from build_plan and is therefore fully known here, before a single row is
    # written, so there is no reason to check them afterwards.
    # -------------------------------------------------------------------------------
    pre_write = Reconciliation(
        counters=plan.counters,
        apply_result=ApplyResult(),
        checksum_matched_approved_source=matches_approved,
        expected_populated_rows=expected_populated_rows,
    )
    if not pre_write.balanced():
        failures = [
            f"  FAIL {name}: {detail}"
            for name, passed, detail in pre_write.balance_checks()
            if not passed
        ]
        raise ImporterError(
            "Reconciliation did NOT balance. Refusing to write anything.\n"
            + "\n".join(failures)
            + "\nNothing has been written."
        )
    print(
        f"Pre-write reconciliation balanced ({len(pre_write.balance_checks())} checks).",
        file=sys.stderr,
    )

    today = datetime.utcnow().strftime("%Y-%m-%d")
    report_slug = MODE_PRODUCTION if mode == MODE_PRODUCTION else "preview"
    report_dir = args.report_dir or Path(
        f"docs/verification/popdam-order-list-{report_slug}-{today}"
    )
    report_dir.mkdir(parents=True, exist_ok=True)

    def write_report(
        applied: ApplyResult,
        *,
        second_run: ApplyResult | None,
        outcome: str,
    ) -> Path:
        """Write the report. NEVER overwrites an existing one.

        Silently replacing a same-day report would destroy the evidence of an earlier
        run in a repo whose approval process is built on these documents.
        """
        path = report_dir / "README.md"
        if path.exists():
            stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
            path = report_dir / f"README-{stamp}.md"
            print(
                f"A report already exists in {report_dir}; writing {path.name} instead "
                "of overwriting it.",
                file=sys.stderr,
            )
        path.write_text(
            render_report(
                reconciliation=Reconciliation(
                    counters=plan.counters,
                    apply_result=applied,
                    checksum_matched_approved_source=matches_approved,
                    expected_populated_rows=expected_populated_rows,
                ),
                checksum=checksum,
                project_ref=project_ref,
                workbook_name=workbook.name,
                run_mode=f"{run_mode} — {outcome}",
                generated_at=datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
                second_run=second_run,
                title=(
                    "PopDAM OrderList production import reconciliation"
                    if mode == MODE_PRODUCTION
                    else "PopDAM OrderList preview import reconciliation"
                ),
                reviewed_commit=reviewed_commit,
            ),
            encoding="utf-8",
        )
        print(f"Report written: {path}", file=sys.stderr)
        return path

    # An abort mid-import must leave a DURABLE record. Previously the exception escaped
    # before the report was written and the only trace was stderr scrollback.
    try:
        first = apply_plan(
            plan,
            gateway,
            batch_size=args.batch_size,
            replace_source=args.replace_source,
            dry_run=args.dry_run,
            allow_replace=allow_replace,
            target_guard=target_guard,
            single_transaction=writing,
        )
    except CommitOutcomeUnknown as error:
        # Distinct from a rollback, deliberately. See CommitOutcomeUnknown.
        write_report(
            ApplyResult(),
            second_run=None,
            outcome=(
                "OUTCOME UNKNOWN. The single COMMIT did not return successfully, so "
                "this run cannot say whether the rows are durable. Establish the real "
                "state before doing anything else; do NOT rerun blindly. "
                f"Cause: {error}"
            ),
        )
        raise
    except Exception as error:
        write_report(
            ApplyResult(),
            second_run=None,
            outcome=(
                "ABORTED before the commit. The whole import ran in ONE transaction "
                f"and was rolled back, so NO rows were written. Cause: {error}"
            ),
        )
        raise

    second: ApplyResult | None = None
    if args.verify_idempotency:
        # The idempotency pass needs the same durable-record guarantee as the first.
        try:
            second = apply_plan(
                plan,
                gateway,
                batch_size=args.batch_size,
                replace_source=args.replace_source,
                dry_run=args.dry_run,
                allow_replace=allow_replace,
                target_guard=target_guard,
                single_transaction=writing,
            )
        except CommitOutcomeUnknown as error:
            write_report(
                first,
                second_run=None,
                outcome=(
                    "FIRST PASS COMPLETED. The idempotency pass then failed to commit, "
                    "so ITS outcome is unknown; the first pass is unaffected. "
                    f"Cause: {error}"
                ),
            )
            raise
        except Exception as error:
            write_report(
                first,
                second_run=None,
                outcome=(
                    "FIRST PASS COMPLETED AND COMMITTED. The idempotency pass then "
                    f"aborted and was rolled back. Cause: {error}"
                ),
            )
            raise
        if second.changed_rows != 0:
            write_report(
                first,
                second_run=second,
                outcome=(
                    "IDEMPOTENCY FAILED: the second run changed "
                    f"{second.changed_rows} business rows."
                ),
            )
            raise ImporterError(
                f"Idempotency check FAILED: the second run changed "
                f"{second.changed_rows} business rows."
            )

    write_report(first, second_run=second, outcome="completed")
    return 0


if __name__ == "__main__":  # pragma: no cover
    try:
        sys.exit(main())
    except ImporterError as error:
        print(f"ERROR: {error}", file=sys.stderr)
        sys.exit(2)
