import json, re
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

rows = json.load(open("laura_answers.json"))
by = {r["ref"]: r for r in rows}

# ---- what is still unresolved ---------------------------------------------
# A002/A004 answered NONE -> settled, dropped.
# A001/A003 (LB, EX): her answer was RIGHT; our property list is missing the code. Our fix, not hers.
ASK_A = ["A005"]
C_MULTI = {"C%03d" % n for n in range(16, 24)}   # answered "DP, XM"
C_OTHER = {"C004", "C006", "C033"}               # blank / sentence / off-list code
ASK_C = sorted(C_MULTI | C_OTHER)
ASK_B = [r["ref"] for r in rows if r["ref"].startswith("B")]

B_DROP = "NOT CHARACTERS - DROP THE ROW"
B_REAL = "REAL CHARACTERS - I LISTED THEM"

wb = Workbook()

# ---------------- sheet 1: how to fill this in ------------------------------
gu = wb.active
gu.title = "How to fill this in"
guide = [
    ("Round 2 - only the questions that are still open", True),
    ("", False),
    ("Thank you - your MG06 codes were right and I kept every one of them.", False),
    ("25 of the 36 character questions are fully closed. Those rows are gone from this sheet.", False),
    ("Lost Boys (LB) and The Exorcist (EX) were also right. The gap is on our side, not yours.", False),
    ("", False),
    ("What to do", True),
    ("Fill in ONLY the yellow columns. Everything else is here for context.", False),
    ("", False),
    ('The "YOUR ANSWER" column is a dropdown. Click the cell, then the arrow, then pick one.', False),
    ("Please do not type into it. If your answer is not in the list, leave it blank and use NOTES.", False),
    ("", False),
    ('"REAL CHARACTER NAMES" only matters if you picked "REAL CHARACTERS - I LISTED THEM".', False),
    ("Separate the names with a semicolon. Example:  Jason; Nate; Shane", False),
    ("", False),
    ("What went wrong last time", True),
    ("On the 154 character rows I asked whether the row names real characters or is just a label.", False),
    ("The answers came back as MG06 property codes, which answer a different question.", False),
    ("That is my fault - my sheet asked for names in a column where every other row wanted a code.", False),
    ("This time the dropdown only allows the two answers I can actually use.", False),
]
for i, (t, bold) in enumerate(guide, 1):
    gu.cell(row=i, column=1, value=t).font = Font(name="Arial", size=11, bold=bold)
gu.column_dimensions["A"].width = 105

# ---------------- sheet 2: the questions ------------------------------------
ws = wb.create_sheet("Open questions")
COLS = ["ref", "question type", "licensor", "style guide", "the row in question",
        "what you answered last time", "why I am asking again", "what I need from you",
        "YOUR ANSWER", "REAL CHARACTER NAMES (only if you picked REAL CHARACTERS)", "NOTES"]
ws.append(COLS)

hdr_fill = PatternFill("solid", fgColor="D9D9D9")
ask_fill = PatternFill("solid", fgColor="FFFF00")
for i in range(1, len(COLS) + 1):
    c = ws.cell(row=1, column=i)
    c.font = Font(name="Arial", size=11, bold=True)
    c.fill = hdr_fill
    c.alignment = Alignment(wrap_text=True, vertical="top")

out = []

for ref in ASK_A:
    r = by[ref]
    out.append(dict(
        ref=ref, qt="Code belongs to the wrong licensor", lic=r["licensor"],
        sg=r["style_guide"], row=r["style_guide"], last=str(r["YOUR_ANSWER"]).strip(),
        why=('You confirmed CC. In our data CC (COCO) sits under a licensor called '
             '"DTR - NO LICENSE", but this style guide is Disney. One code cannot belong '
             'to two licensors.'),
        need="Confirm CC anyway, or give the Disney code, or say NONE if POP never produced it.",
        dv="CC,NONE"))

for ref in ASK_B:
    r = by[ref]
    comps = re.sub(r"^.*?real characters:\s*", "", r["the_problem"], flags=re.I).rstrip(".")
    last = str(r["YOUR_ANSWER"]).strip()
    if last.upper() == "NONE":
        why = ('You wrote NONE. I could not tell whether that means "drop this row" or '
               '"no property code".')
    else:
        why = ('You answered "%s", which is an MG06 property code. That tells me the property, '
               "not whether these are real characters." % last)
    out.append(dict(
        ref=ref, qt="One row names several characters", lic=r["licensor"],
        sg=r["style_guide"], row=r["character"], last=last, why=why,
        need="Are these real characters we should add, or just a label? Unknown names: %s" % comps,
        dv="%s,%s" % (B_DROP, B_REAL)))

for ref in ASK_C:
    r = by[ref]
    opts = [o.strip() for o in str(r["options"]).split("/") if o.strip()]
    last = str(r["YOUR_ANSWER"]).strip()
    if ref in C_MULTI:
        why = 'You answered "%s" - two codes. A character can only belong to ONE property.' % last
    elif ref == "C004":
        why = "This one was left blank."
    elif ref == "C006":
        why = ('You answered "%s", which lists three characters instead of picking one code '
               "for this row." % last)
    else:
        why = ('You answered "%s". That code was not one of the choices and does not exist in '
               "our property list." % last)
    out.append(dict(
        ref=ref, qt="Character sits under two or more MG06 codes", lic=r["licensor"],
        sg=r["style_guide"], row=r["character"], last=last, why=why,
        need="Pick exactly ONE code.",
        dv=",".join(opts + ["NONE"])))

for o in out:
    ws.append([o["ref"], o["qt"], o["lic"], o["sg"], o["row"], o["last"], o["why"], o["need"], "", "", ""])

for i, o in enumerate(out, start=2):
    dv = DataValidation(type="list", formula1='"%s"' % o["dv"], allow_blank=True, showDropDown=False)
    dv.errorTitle = "Pick from the list"
    dv.error = "Please pick one of the listed answers. If none fit, leave it blank and use NOTES."
    dv.promptTitle = "Choose one"
    dv.prompt = "Click the arrow and pick one."
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=i, column=9))

for i in range(2, len(out) + 2):
    for col in (9, 10, 11):
        ws.cell(row=i, column=col).fill = ask_fill
    for col in range(1, 12):
        c = ws.cell(row=i, column=col)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

for i, w in enumerate([8, 26, 14, 26, 34, 16, 46, 44, 30, 32, 22], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:K%d" % (len(out) + 1)

path = (r"C:\repos\shared-db\docs\verification\character-identity-rules-20260728"
        r"\licensing-questions-for-laura-round2-20260731.xlsx")
wb.save(path)
print("wrote", path)
print("rows:", len(out), "| A:", len(ASK_A), "B:", len(ASK_B), "C:", len(ASK_C))
print("removed as already resolved:", len(rows) - len(out))
