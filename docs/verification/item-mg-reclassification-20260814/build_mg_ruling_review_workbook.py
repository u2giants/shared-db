"""Build the reviewer workbook for the depth-3 signatures that block the most historical items.

Every historical item that fails to reach MG01+MG02+MG03 is blocked by one product
signature. The blockage is heavily concentrated: a small number of signatures account
for most of the shortfall, and each one needs a single business ruling rather than a
change to the matching code.

This workbook presents those signatures in blocked-row order and, for each competing
merchandise-group candidate, lists every post-change item that voted for it so the
ruling can be spot checked against real descriptions.

No database data is read or written. Row-level sheets are generated into the private
output folder and are not committed.
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import asdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import hierarchical_item_taxonomy as H

DEPTH = 3
ITEM_COLUMNS = ["Company", "Division", "Item #", "Item Desc", "CreatedTime", "MG01", "MG02", "MG03"]
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
RULING_FILL = PatternFill("solid", fgColor="FFF2CC")


def blocking_reason(evidence: Counter) -> str:
    """Explain why choose_at_level rejected this signature, in reviewer language."""
    if not evidence:
        return "No post-change item has ever used this signature"
    ranked = evidence.most_common()
    chosen, count = ranked[0]
    total = sum(evidence.values())
    runner = ranked[1][1] if len(ranked) > 1 else 0
    if total < 2:
        return f"Only {total} post-change item exists, so no majority can be established"
    share = count / total
    if share < 0.80:
        return f"Top candidate {chosen} holds only {share:.0%} of the evidence; 80% is required"
    if runner and count < runner * 1.5:
        return f"Top candidate {chosen} does not lead the runner-up by the required 1.5x"
    return "Not blocked"


def signature_of(row: pd.Series) -> str:
    return "|".join(H.normalize(row[field]) for field in ("Product Type", "Construction Shape", "Treatment"))


def parse_source(source: Path, reference: Path | None, dictionary_path: Path) -> pd.DataFrame:
    data = pd.read_csv(source, dtype=str, encoding="cp1252").fillna("")
    data["Created Date"] = pd.to_datetime(data["CreatedTime"], errors="coerce")
    if data["Created Date"].isna().any():
        raise ValueError(f"{int(data['Created Date'].isna().sum())} CreatedTime values could not be read")
    dictionary = H.load_product_dictionary(dictionary_path)
    licensors, properties = H.load_name_lexicons(reference)
    parsed = [H.parse_description(value, licensors, properties, dictionary) for value in data["Item Desc"]]
    for field in H.ParsedDescription.__dataclass_fields__:
        data[field.replace("_", " ").title()] = [asdict(value)[field] for value in parsed]
    data = data.rename(columns={"MerchGroup01": "MG01", "MerchGroup02": "MG02", "MerchGroup03": "MG03"})
    data["Description Usable for Product Matching"] = data["Item Desc"].map(
        lambda value: "Yes" if H.usable_description(value) else "No"
    )
    return data


def build(source: Path, output: Path, reference: Path | None, dictionary_path: Path, top: int) -> dict:
    data = parse_source(source, reference, dictionary_path)
    post = data[data["Created Date"].ge(H.CUTOFF)].copy()
    pre = data[data["Created Date"].lt(H.CUTOFF)].copy()
    _, reverse = H.build_associations(post)

    accepted = pre[pre["Dictionary Status"].isin(["accepted", "alias"]) & pre["Product Type"].ne("")].copy()
    accepted["Signature"] = accepted.apply(signature_of, axis=1)
    blocked = accepted[[H.choose_at_level(value, DEPTH, reverse[DEPTH]) is None for value in accepted["Signature"]]]

    counts = Counter(blocked["Signature"])
    shortfall = sum(counts.values())
    ranked_signatures = [name for name, _ in counts.most_common(top)]

    post["Signature"] = post.apply(signature_of, axis=1)

    decisions, examples, affected = [], [], []
    running = 0
    for rank, signature in enumerate(ranked_signatures, start=1):
        evidence = Counter(reverse[DEPTH].get(signature, {}))
        ranked = evidence.most_common()
        total = sum(evidence.values())
        running += counts[signature]
        product, construction, treatment = (signature.split("|") + ["", "", ""])[:3]
        pairs = {key.rsplit("|", 1)[0] for key in evidence}
        decisions.append({
            "Rank": rank,
            "Historical Items Blocked": counts[signature],
            "Share of Total Shortfall": round(counts[signature] / shortfall, 4),
            "Cumulative Share of Shortfall": round(running / shortfall, 4),
            "Signature": signature,
            "Product Type": product,
            "Construction / Shape": construction,
            "Treatment Stated in Description": treatment or "(none stated)",
            "Why It Is Blocked": blocking_reason(evidence),
            "Post-Change Items Voting": total,
            "Distinct Candidates": len(ranked),
            "Candidate 1": ranked[0][0] if ranked else "",
            "Candidate 1 Items": ranked[0][1] if ranked else 0,
            "Candidate 1 Share": round(ranked[0][1] / total, 4) if total else 0,
            "Candidate 2": ranked[1][0] if len(ranked) > 1 else "",
            "Candidate 2 Items": ranked[1][1] if len(ranked) > 1 else 0,
            "Candidate 2 Share": round(ranked[1][1] / total, 4) if len(ranked) > 1 and total else 0,
            "All Candidates": "; ".join(f"{key}: {n}" for key, n in ranked),
            "MG01+MG02 Agreed by All Candidates": "Yes" if len(pairs) == 1 else "No",
            "RULING - MG Key to Apply": "",
            "RULING - Reviewer": "",
            "RULING - Notes": "",
        })

        supporters = post[post["Signature"].eq(signature)].copy()
        supporters["Candidate MG Key"] = supporters.apply(lambda row: H.hierarchy_key(row, DEPTH), axis=1)
        supporters = supporters[supporters["Candidate MG Key"].isin(evidence)]
        order = {key: index for index, (key, _) in enumerate(ranked)}
        supporters["_order"] = supporters["Candidate MG Key"].map(order)
        for _, row in supporters.sort_values(["_order", "Item #"]).iterrows():
            examples.append({
                "Rank": rank,
                "Signature": signature,
                "Candidate MG Key": row["Candidate MG Key"],
                "Candidate Items": evidence[row["Candidate MG Key"]],
                "Candidate Share": round(evidence[row["Candidate MG Key"]] / total, 4),
                **{column: row[column] for column in ITEM_COLUMNS},
            })

        for _, row in blocked[blocked["Signature"].eq(signature)].sort_values("Item #").iterrows():
            affected.append({
                "Rank": rank,
                "Signature": signature,
                "Company": row["Company"], "Division": row["Division"], "Item #": row["Item #"],
                "Item Desc": row["Item Desc"], "CreatedTime": row["CreatedTime"],
                "Stored MG01": row["MG01"], "Stored MG02": row["MG02"], "Stored MG03": row["MG03"],
                "Note": "Stored values are shown for context only and are never used as evidence",
            })

    readme = pd.DataFrame({"How to use this workbook": [
        f"Every row on Decisions is one product signature that blocks historical items from a full "
        f"MG01+MG02+MG03 result. The top {top} signatures shown here account for "
        f"{running / shortfall:.0%} of the entire {shortfall:,}-item shortfall.",
        "Read Decisions in Rank order. Rank 1 blocks the most items, so it is worth the most review time.",
        "For each signature, Candidate Examples lists every post-change item that was assigned each "
        "competing MG key, with its full description. Filter by Signature, then by Candidate MG Key, to "
        "compare the real items behind each candidate side by side and spot check which one is right.",
        "The Historical Items Awaiting sheet lists the pre-May-14-2025 items that each ruling would resolve.",
        "Record your decision in the three RULING columns on Decisions. A ruling may name one candidate, "
        "split the signature into finer treatments, or state that the description cannot determine MG03.",
        "'(none stated)' in Treatment means the description names no treatment at all. MG03 encodes a "
        "treatment, so these may be genuinely undecidable from the description; saying so is a valid ruling.",
        "Stored merchandise-group values on historical items are shown for context only. They were created "
        "under the old method and are never used as evidence.",
        "No database data was read or changed to produce this workbook.",
    ]})

    frames = {
        "Readme": readme,
        "Decisions": pd.DataFrame(decisions),
        "Candidate Examples": pd.DataFrame(examples),
        "Historical Items Awaiting": pd.DataFrame(affected),
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in frames.items():
            frame.to_excel(writer, sheet_name=name, index=False)
        for name, frame in frames.items():
            style_sheet(writer.sheets[name], frame)

    return {
        "signatures_reviewed": len(ranked_signatures),
        "total_shortfall_rows": shortfall,
        "rows_covered_by_this_workbook": running,
        "share_of_shortfall_covered": round(running / shortfall, 4),
        "candidate_example_rows": len(examples),
        "workbook": str(output),
    }


def style_sheet(sheet, frame: pd.DataFrame) -> None:
    widths = {"Item Desc": 62, "All Candidates": 58, "Why It Is Blocked": 52, "Signature": 34,
              "How to use this workbook": 118, "Note": 44}
    for index, column in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        longest = max([len(str(column))] + [len(str(value)) for value in frame[column].head(400)]) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(widths.get(column, longest), 62)
        if str(column).startswith("RULING"):
            for row in range(2, len(frame) + 2):
                sheet.cell(row=row, column=index).fill = RULING_FILL
    sheet.freeze_panes = "A2"
    if len(frame):
        sheet.auto_filter.ref = sheet.dimensions
    if "How to use this workbook" in frame.columns:
        for row in range(2, len(frame) + 2):
            sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            sheet.row_dimensions[row].height = 46


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--reference", type=Path)
    parser.add_argument("--dictionary", type=Path, default=Path(__file__).with_name("product_type_dictionary.csv"))
    parser.add_argument("--top", type=int, default=25)
    args = parser.parse_args()
    print(json.dumps(build(args.source, args.output, args.reference, args.dictionary, args.top), indent=2))


if __name__ == "__main__":
    main()
